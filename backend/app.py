"""IAM SoD Backend (Phase 9, Bau-Schritt 1) — Runner-as-API.

Orchestriert die vorhandenen cypher/-Dateien ueber den Neo4j-Treiber (apoc.cypher.runFile),
loest Profile aus config/analysis_profiles.json auf und faehrt Materialisierung+Auswertung als
asynchronen Job. Plattformunabhaengig im Container; ersetzt die PowerShell-Runner durch HTTP.
Bewusst MVP: In-Memory-Jobs (Single-Instance), Findings bleiben im Graph (kein eigener Store).
"""
import os
import re
import io
import csv
import json
import time
import uuid
import shutil
import zipfile
import tempfile
import datetime
import threading
from pathlib import Path

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel
from neo4j import GraphDatabase

import convert

NEO4J_URI = os.environ.get("NEO4J_URI", "bolt://neo4j:7687")
NEO4J_USER = os.environ.get("NEO4J_USER", "neo4j")
NEO4J_PASSWORD = os.environ["NEO4J_PASSWORD"]
CONFIG_DIR = Path(os.environ.get("CONFIG_DIR", "/app/config"))
RULES_DIR = Path(os.environ.get("RULES_DIR", "/app/rules"))
CHECKS_DIR = Path(os.environ.get("CHECKS_DIR", "/app/checks"))
CHECK_AREAS = {"user": ["A", "B", "C", "D", "E"], "role": ["R"], "import": ["I"]}
CYPHER_DIR = Path(os.environ.get("CYPHER_DIR", "/app/cypher"))
LOAD_DIR = Path(os.environ.get("LOAD_DIR", "/app/load"))
MIGRATIONS_DIR = Path(os.environ.get("MIGRATIONS_DIR", "/app/migrations"))
DATA_DIR = Path(os.environ.get("DATA_DIR", "/app/data/import"))
BACKUP_DIR = Path(os.environ.get("BACKUP_DIR", "/app/backups"))
RUN_BACKUP_DIR = BACKUP_DIR / "runs"
LOG_DIR = Path(os.environ.get("LOG_DIR", "/app/data/logs"))
JOB_ERROR_LOG = LOG_DIR / "job_errors.jsonl"
FRONTEND_DIR = Path(os.environ.get("FRONTEND_DIR", "/app/frontend"))
DEFAULT_LANG = [c.strip() for c in os.environ.get("IMPORT_LANG", "DE,DEU,D").split(",") if c.strip()]

driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASSWORD))
app = FastAPI(title="IAM SoD Backend", version="0.1.0")
jobs: dict[str, dict] = {}  # jobId -> status


# --- Fehlerprotokoll (Job-Fehler) -------------------------------------------------------
# Persistiert fehlgeschlagene Jobs ueber Neustarts hinweg (jobs-Dict ist nur In-Memory).
# JSONL unter data/logs (Bind-Mount, ueberlebt Container-Neustart) statt Graph: rein operative
# Nachvollziehbarkeit, keine fachliche Ableitung (AE-10 betrifft die Findings-Schicht, nicht das).
def _log_job_error(job_id: str, message: str) -> None:
    entry = {
        "ts": datetime.datetime.now().isoformat(timespec="seconds"),
        "jobId": job_id,
        "kind": jobs.get(job_id, {}).get("kind"),
        "request": jobs.get(job_id, {}).get("request"),
        "message": message,
    }
    try:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        with open(JOB_ERROR_LOG, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except OSError:
        pass


@app.middleware("http")
async def no_cache_frontend(request, call_next):
    """Statische UI aendert sich oft (laufende Entwicklung) -> Browser sollen IMMER
    revalidieren (ETag/Last-Modified -> billiges 304 bei unveraendertem Inhalt) statt eine
    alte index.html ungeprueft aus dem Cache zu zeigen."""
    response = await call_next(request)
    if request.url.path == "/" or request.url.path.endswith(".html"):
        response.headers["Cache-Control"] = "no-cache"
    return response


ORG_PROFILES_VENDOR_PATH = CONFIG_DIR / "analysis_profiles.json"
ORG_PROFILES_CUSTOM_PATH = CONFIG_DIR / "analysis_profiles.custom.json"
PROTECTED_ORG_PROFILES = {"standard", "uebergreifend"}


def ensure_custom_org_profiles_file() -> Path:
    if not ORG_PROFILES_CUSTOM_PATH.is_file():
        ORG_PROFILES_CUSTOM_PATH.write_text("[]", encoding="utf-8")
    return ORG_PROFILES_CUSTOM_PATH


def profiles() -> dict:
    """Liest die Vendor-Profile UND merged die Org-Profile-Overlay-Datei
    (analysis_profiles.custom.json, vom Admin-Editor "Org-Varianten" geschrieben) hinein —
    analog zum Ruleset-Overlay-Mechanismus (queries.custom.json/sod_rules.custom.json), aber
    global statt pro Ruleset. Overlay-Namen, die mit einem Vendor-Profil kollidieren, werden
    beim Schreiben verhindert (s. admin_create_org_profile), hier nur defensiv gefiltert."""
    cfg = json.loads(ORG_PROFILES_VENDOR_PATH.read_text(encoding="utf-8"))
    vendor_names = {p["name"] for p in cfg["profiles"]}
    custom = _load_json_list(ORG_PROFILES_CUSTOM_PATH)
    cfg["profiles"] = cfg["profiles"] + [p for p in custom if p["name"] not in vendor_names]
    return cfg


def list_rulesets() -> list[dict]:
    """Verfuegbare Rulesets aus rules/*/ruleset.json (id + Anzeigename)."""
    out = []
    if RULES_DIR.exists():
        for d in sorted(RULES_DIR.iterdir()):
            rj = d / "ruleset.json"
            if rj.is_file():
                meta = json.loads(rj.read_text(encoding="utf-8"))
                if meta.get("ruleset"):
                    out.append({"id": meta["ruleset"], "name": meta.get("name", meta["ruleset"]), "dir": d.name})
    return out


def ruleset_dir(ruleset: str):
    for r in list_rulesets():
        if r["id"] == ruleset:
            return r["dir"]
    return None


# --- Einzelfilter-Editor (Query-Metadaten) ----------------------------------------------
# Schreibt NIE in die Vendor-Datei (queries.json) — Edits/abgeleitete Queries landen in einem
# Overlay (queries.custom.json) je Ruleset-Ordner, das load_ruleset.cypher zusaetzlich einliest
# (siehe dort: coalesce-Merge, Overlay gewinnt bei gesetzten Feldern). Bewusst nur Metadaten in
# v1 (description/criticality/module/queryType/disregardTcode) — authorizations/transactions
# werden bei "Ableiten" 1:1 von der Quelle kopiert, aber hier nicht editiert.
def _ruleset_paths(ruleset: str) -> tuple[Path, Path]:
    rdir = ruleset_dir(ruleset)
    if not rdir:
        raise HTTPException(404, f"Ruleset '{ruleset}' nicht gefunden")
    base = RULES_DIR / rdir
    return base / "queries.json", base / "queries.custom.json"


def _load_json_list(path: Path) -> list[dict]:
    return json.loads(path.read_text(encoding="utf-8")) if path.is_file() else []


def ensure_custom_queries_file(ruleset: str) -> Path:
    _, custom_path = _ruleset_paths(ruleset)
    if not custom_path.is_file():
        custom_path.write_text("[]", encoding="utf-8")
    return custom_path


def _merged_queries(ruleset: str) -> tuple[dict[str, dict], set[str]]:
    """Vendor-Queries + Overlay, Overlay-Felder gewinnen je id. Gibt (id -> effektive Query,
    Menge der STRUKTURELL eigenen ids) zurueck — 'eigen' heisst hier: komplett neue/abgeleitete
    Query (kein Vendor-Gegenstueck) ODER der Aufbau (authorizations/transactions) wurde
    ueberschrieben. Reine Metadaten-Ergaenzungen (Kurzbezeichnung, Risiko, Controls, ...) auf
    einer bestehenden Vendor-Query zaehlen NICHT als 'eigen' (s. Nutzerfeedback: das ist nur
    eine Info-Ergaenzung, keine inhaltliche Aenderung der Query)."""
    vendor_path, custom_path = _ruleset_paths(ruleset)
    merged = {q["query"]: dict(q) for q in _load_json_list(vendor_path)}
    custom_ids = set()
    for c in _load_json_list(custom_path):
        qid = c["query"]
        if qid not in merged or "authorizations" in c or "transactions" in c:
            custom_ids.add(qid)
        if qid in merged:
            merged[qid] = {**merged[qid], **{k: v for k, v in c.items() if v is not None}}
        else:
            merged[qid] = c
    return merged, custom_ids


def reload_ruleset(ruleset: str):
    rdir = ruleset_dir(ruleset)
    ensure_custom_queries_file(ruleset)    # beide Overlay-Dateien muessen existieren (apoc.load.json)
    ensure_custom_sodrules_file(ruleset)
    with driver.session() as s:
        run_file(s, "ruleset/load_ruleset.cypher", {"dir": rdir, "ruleset": ruleset})


def jsonable(v):
    """Neo4j-Temporaltypen (Date/DateTime/Time) -> ISO-String; Container rekursiv."""
    if hasattr(v, "isoformat"):
        return v.isoformat()
    if isinstance(v, dict):
        return {k: jsonable(x) for k, x in v.items()}
    if isinstance(v, (list, tuple)):
        return [jsonable(x) for x in v]
    return v


def split_statements(text: str) -> list[str]:
    """Cypher-Datei in Einzel-Statements zerlegen (wie cypher-shell `-f`): volle //-Kommentarzeilen
    entfernen, an ';' trennen. apoc.cypher.runFile waere apoc-extended; der Container hat nur
    apoc-core -> wir fahren die Statements direkt ueber den Treiber (je Statement Auto-Commit)."""
    lines = [ln for ln in text.splitlines() if not ln.lstrip().startswith("//")]
    stmts = [s.strip() for s in "\n".join(lines).split(";")]
    return [s for s in stmts if s]


def run_cypher_path(session, path: Path, params: dict) -> dict:
    """Eine .cypher-Datei ausfuehren; gibt aggregierte Summary-Zaehler zurueck."""
    totals = {"nodes_created": 0, "relationships_created": 0, "properties_set": 0}
    for stmt in split_statements(path.read_text(encoding="utf-8")):
        c = session.run(stmt, **params).consume().counters
        totals["nodes_created"] += c.nodes_created
        totals["relationships_created"] += c.relationships_created
        totals["properties_set"] += c.properties_set
    return totals


def run_file(session, rel_path: str, params: dict):
    run_cypher_path(session, CYPHER_DIR / rel_path, params)


class RunReq(BaseModel):
    ruleset: str = "kpmg_r3"
    dataset: str
    # Kein asOf mehr hier: der Stichtag ist eine Eigenschaft des Datasets (= Downloaddatum der
    # SAP-Extrakte), keine Lauf-/Filter-Eingabe. Wird ueber _dataset_asof() aufgeloest; Korrektur
    # nur global ueber PUT /datasets/{id}/asof.
    userTypeProfile: str = "all"
    orgProfile: str = "standard"
    # "all" (Default) materialisiert MATCHES fuer JEDE Query des Rulesets -- auch Einzelfilter,
    # die in keiner SoD-Regel als Klausel verbaut sind. "sodOnly" beschraenkt die materialize-Phase
    # auf die tatsaechlich von den SoD-Regeln benoetigten Queries (schneller, aber die Einzelfilter-
    # Ergebnisse/Uebersicht zeigen dann nur diese Teilmenge) -- Nutzer-Wunsch: Standard ist "alles
    # rechnen", Einschraenkung ist die bewusste Ausnahme (Ruleset-abhaengig, z. B. anders beim
    # CSI-Filterset).
    queryScope: str = "all"
    sleepDays: int | None = None
    minCriticalityRank: int = 0
    sodRules: list[str] = []
    # Explizite Einzelfilter-Auswahl (Katalog-Auswahl, Assistent Schritt ③) -- nicht leer:
    # materialisiert NUR diese Queries, unabhaengig von queryScope/sodRules (s.
    # materialize_matches_candidates.cypher). Leer (Default) = bisheriges Verhalten.
    queryIds: list[str] = []
    # False = Can-Do-Modus (Katalog-Auswahl mit reinen Einzelfiltern, ohne SoD-Regeln): die
    # evaluate-Phase legt weiterhin den (:Run)-Knoten an, ueberspringt aber die eigentliche
    # SoD-Regel-Auswertung + Explain komplett (s. _run_one).
    evaluateSod: bool = True
    runId: str | None = None
    title: str | None = None        # menschenlesbarer Name der Variante (z. B. "Übergreifend");
                                     # leer -> runId als Fallback (siehe do_run)
    skipRulesetLoad: bool = True
    skipMaterialize: bool = False
    skipExplain: bool = False        # Evidenz (VIA_ROLE/VIA_PROFILE + intra/inter) -- seit
                                     # Evidenz-Perf (GRANTS-Kante + Checkpoint-Throttling +
                                     # explain_sod_finalize-Fix, s. ROADMAP.md) Default AN;
                                     # abwaehlbar fuer schnellere Laeufe, sonst per Formular oder
                                     # POST /runs/{id}/explain nachtraeglich anforderbar
    resume: bool = False             # Ab dem gespeicherten Lauf-Checkpoint weitermachen (_run_state.json);
                                     # alle anderen Felder werden dann ignoriert -> die Original-Parameter
                                     # aus dem Checkpoint gelten (siehe do_run()).


def _slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def _humanize_profile_name(name: str) -> str:
    """Kurzer, lesbarer Lauf-Titel aus dem Org-Profil-Namen (z. B. 'bukrs-1000-und-2000' ->
    'Bukrs 1000 Und 2000') — analog zu defaultRunTitle() im Frontend (frontend/index.html), damit
    Einzel- und Batch-Laeufe denselben Titel-Stil haben (die volle Beschreibung waere als
    Run-Listen-Label zu lang)."""
    return " ".join(w[:1].upper() + w[1:] for w in re.split(r"[-_]+", name) if w)


def _checkpoint_path(dataset: str, name: str) -> Path:
    return DATA_DIR / dataset / name

def _read_checkpoint(path: Path) -> dict | None:
    try:
        return json.loads(path.read_text(encoding="utf-8")) if path.is_file() else None
    except Exception:
        return None

def _write_checkpoint(path: Path, state: dict) -> None:
    state = {**state, "updatedAt": datetime.datetime.now().isoformat(timespec="seconds")}
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

def _clear_checkpoint(path: Path) -> None:
    if path.is_file():
        path.unlink()


def _run_candidates(s, rel_path: str, params: dict) -> list[str]:
    """Fuehrt eine Kandidaten-Datei aus (ein einzelnes RETURN id-Statement) und gibt die
    Einheiten-IDs zurueck, die _run_phase() Schritt fuer Schritt abarbeitet."""
    stmt = split_statements((CYPHER_DIR / rel_path).read_text(encoding="utf-8"))[0]
    return [r["id"] for r in s.run(stmt, **params)]


def _run_phase(s, job_id: str, *, phase: str, step_label: str, state_path: Path,
               reset_rel_path: str | None, candidates_rel_path: str, one_rel_path: str,
               unit_param: str, params: dict, resume_units: set[str],
               checkpoint_extra: dict) -> None:
    """Eine Phase (materialize/evaluate/explain-PROVIDES) als 'Reset einmalig -> Kandidaten
    ermitteln -> pro Einheit' statt als ein einziger, nicht unterbrechbarer Aufruf. Bei
    resume_units (nicht leer): reset_rel_path wird uebersprungen (sonst waeren bereits fertige
    Einheiten wieder weg) und bereits erledigte Einheiten werden nicht erneut gerechnet.
    Checkpoint (data/<dataset>/_run_state.json bzw. das vom Aufrufer uebergebene state_path) wird
    zeitgesteuert geschrieben (hoechstens alle CHECKPOINT_MIN_INTERVAL Sekunden, plus garantiert
    nach der letzten Einheit) statt nach JEDER Einheit -- ein volles Neuschreiben der wachsenden
    completedUnits-Liste bei jeder Einheit ist O(n^2) und wurde bei mehreren tausend Einheiten
    (z. B. Evidenz-Akteure) zum dominanten Zeitfresser, weit vor der eigentlichen Cypher-Query
    (s. Evidenz-Perf-Benchmark). Ein Absturz verliert dadurch hoechstens die letzten paar Sekunden
    bereits erledigter Einheiten -- unkritisch, da jede Einheit idempotent per MERGE schreibt und
    beim Resume einfach nochmal (billig) laeuft."""
    if not resume_units and reset_rel_path:
        run_file(s, reset_rel_path, params)

    candidates = _run_candidates(s, candidates_rel_path, params)
    total = len(candidates)
    completed = list(resume_units)
    completed_set = set(completed)

    def _persist():
        _write_checkpoint(state_path, {**checkpoint_extra, "phase": phase,
                                       "completedUnits": completed, "unitTotal": total})

    jobs[job_id].update(step=step_label, stepNum=len(completed), stepTotal=total)
    CHECKPOINT_MIN_INTERVAL = 2.0   # Sekunden
    last_persist = time.monotonic()
    for unit_id in candidates:
        if unit_id in completed_set:
            continue
        _check_cancel(job_id)
        run_file(s, one_rel_path, {**params, unit_param: unit_id})
        completed.append(unit_id)
        completed_set.add(unit_id)
        jobs[job_id].update(stepNum=len(completed), stepTotal=total)
        now = time.monotonic()
        if now - last_persist >= CHECKPOINT_MIN_INTERVAL:
            _persist()
            last_persist = now
    jobs[job_id].pop("stepNum", None)
    jobs[job_id].pop("stepTotal", None)
    if completed:
        _persist()   # garantierter Abschluss-Checkpoint, auch wenn das Intervall nicht erreicht wurde


def _explain_one(s, job_id: str, *, ruleset: str, dataset: str, as_of, run_id: str,
                 state_path: Path, checkpoint_extra: dict, resume_units: set[str],
                 step_prefix: str = "") -> None:
    """Evidenz (VIA_ROLE/VIA_PROFILE + conflictType) fuer einen Lauf berechnen -- PROVIDES
    (teuer, pro Akteur) ueber _run_phase() mit Fortschritt/Resume, der Abschluss
    (explain_sod_finalize.cypher, bereits auf die Findings dieses Laufs begrenzt und damit
    schnell) danach als ein Aufruf. Von _run_one() (explain-Phase eines Laufs) UND do_explain()
    (Ribbon-Button "Evidenz nachrechnen") gemeinsam genutzt."""
    base = {"ruleset": ruleset, "dataset": dataset, "asOf": as_of, "runId": run_id}
    _run_phase(s, job_id, phase="explain", step_label=step_prefix + "explain",
               state_path=state_path, reset_rel_path=None,
               candidates_rel_path="sod/explain_sod_candidates.cypher",
               one_rel_path="sod/explain_sod_one.cypher", unit_param="actorId",
               params=base, resume_units=resume_units, checkpoint_extra=checkpoint_extra)
    _check_cancel(job_id)
    run_file(s, "sod/explain_sod_finalize.cypher", base)


_RUN_PHASE_ORDER = ["ruleset", "materialize", "evaluate", "explain"]


def _run_one(s, job_id: str, cfg: dict, *, ruleset: str, dataset: str, user_type_profile: str,
             org_profile: str, sleep_days: int, min_criticality_rank: int, sod_rules: list[str],
             query_scope: str, query_ids: list[str], evaluate_sod: bool, run_id: str, title: str,
             skip_ruleset_load: bool, skip_materialize: bool,
             skip_explain: bool, step_prefix: str = "", checkpoint_base: dict | None = None,
             resume_phase: str | None = None, resume_units: list[str] | None = None) -> dict:
    """Ein einzelner Lauf ([ruleset] -> materialize -> evaluate -> [explain] -> Ergebnis-
    Zaehlung). Von do_run() (1 Variante) UND do_run_batch() (mehrere Varianten, gemeinsame
    Session) genutzt — step_prefix erlaubt dem Batch, den Fortschritt je Variante anzuzeigen
    (z. B. "Variante 2/3 (uebergreifend): materialize"). materialize/evaluate/explain laufen
    ueber _run_phase()/_explain_one() mit Fortschritt + Checkpoint (data/<dataset>/_run_state.json,
    checkpoint_base traegt die dataset-weiten Zusatzfelder wie runId/params/Batch-Kontext).
    resume_phase/resume_units setzen an einem frueheren Abbruch fort: Phasen VOR resume_phase
    gelten als bereits abgeschlossen (werden uebersprungen), resume_phase selbst macht bei den
    darin bereits erledigten Einheiten weiter, alles danach laeuft frisch."""
    utp = next((p for p in cfg["userTypeProfiles"] if p["name"] == user_type_profile), None)
    if not utp:
        raise ValueError(f"userTypeProfile '{user_type_profile}' unbekannt")
    op = next((p for p in cfg["profiles"] if p["name"] == org_profile), None)
    if not op:
        raise ValueError(f"orgProfile '{org_profile}' unbekannt")
    org_mode = op["org"].get("mode", "ignoreOrg")
    org_filters = op["org"].get("filters", {}) if org_mode == "filtered" else {}
    org = {"orgMode": org_mode, "orgFilters": org_filters}

    as_of = _dataset_asof(s, dataset)
    base = {"ruleset": ruleset, "dataset": dataset, "asOf": as_of, "runId": run_id}
    state_path = _checkpoint_path(dataset, "_run_state.json")
    checkpoint_extra = checkpoint_base or {}

    enabled = {"ruleset": not skip_ruleset_load, "materialize": not skip_materialize,
               "evaluate": True, "explain": not skip_explain and evaluate_sod}
    reached = resume_phase is None
    resume_units_set = set(resume_units or [])

    for phase in _RUN_PHASE_ORDER:
        if not enabled[phase]:
            continue
        if not reached:
            if phase == resume_phase:
                reached = True
            else:
                continue   # vor dem Wiederaufnahmepunkt liegende Phase -> schon abgeschlossen

        this_phase_resume = resume_units_set if phase == resume_phase else set()

        if phase == "ruleset":
            _check_cancel(job_id)
            rdir = ruleset_dir(ruleset)
            if not rdir:
                raise ValueError(f"Ruleset-Ordner fuer '{ruleset}' nicht gefunden")
            jobs[job_id]["step"] = step_prefix + "ruleset"
            ensure_custom_queries_file(ruleset)    # queries.custom.json muss existieren (apoc.load.json)
            ensure_custom_sodrules_file(ruleset)   # sod_rules.custom.json ebenso
            run_file(s, "ruleset/load_ruleset.cypher", {"dir": rdir, "ruleset": ruleset})

        elif phase == "materialize":
            _run_phase(s, job_id, phase="materialize", step_label=step_prefix + "materialize",
                       state_path=state_path, reset_rel_path="sod/materialize_matches_reset.cypher",
                       candidates_rel_path="sod/materialize_matches_candidates.cypher",
                       one_rel_path="sod/materialize_matches_one.cypher", unit_param="qid",
                       params={**base, **org, "queryScope": query_scope, "sodRules": sod_rules,
                               "queryIds": query_ids},
                       resume_units=this_phase_resume, checkpoint_extra=checkpoint_extra)

        elif phase == "evaluate":
            _check_cancel(job_id)
            jobs[job_id]["step"] = step_prefix + "evaluate"
            eval_params = {
                **base, **org,
                "userTypes": list(utp.get("userTypes", [])),
                "excludeLocked": bool(utp.get("excludeLocked", False)),
                "sleepDays": sleep_days,
                "minCriticalityRank": min_criticality_rank,
                "sodRules": sod_rules,
                "queryIds": query_ids,
                "queryScope": query_scope,
                "title": title,
            }
            if not this_phase_resume:
                # Legt/aktualisiert den (:Run)-Knoten IMMER an (Ergebnis-Views wie /queries/summary
                # brauchen ihn) -- auch im Can-Do-Modus (evaluate_sod=False), dort dann ohne
                # anschliessenden Regel-Loop.
                run_file(s, "sod/evaluate_sod_init.cypher", eval_params)
            if evaluate_sod:
                _run_phase(s, job_id, phase="evaluate", step_label=step_prefix + "evaluate",
                           state_path=state_path, reset_rel_path=None,
                           candidates_rel_path="sod/evaluate_sod_candidates.cypher",
                           one_rel_path="sod/evaluate_sod_one.cypher", unit_param="ruleId",
                           params=eval_params, resume_units=this_phase_resume,
                           checkpoint_extra=checkpoint_extra)

        elif phase == "explain":
            _explain_one(s, job_id, ruleset=ruleset, dataset=dataset, as_of=as_of, run_id=run_id,
                        state_path=state_path, checkpoint_extra=checkpoint_extra,
                        resume_units=this_phase_resume, step_prefix=step_prefix)

    rec = s.run(
        "MATCH (f:SoDConflict {runId:$r}) "
        "RETURN count(f) AS findings, count(DISTINCT f.ruleId) AS rules, "
        "sum(CASE WHEN f.userSleeping THEN 1 ELSE 0 END) AS sleeping, "
        "sum(CASE WHEN f.conflictType='intra' THEN 1 ELSE 0 END) AS intra", r=run_id,
    ).single()
    return {"runId": run_id, "title": title, "findings": rec["findings"], "rules": rec["rules"],
            "sleeping": rec["sleeping"], "intra": rec["intra"]}


def do_run(job_id: str, req: RunReq):
    try:
        cfg = profiles()
        resume_state = _read_checkpoint(_checkpoint_path(req.dataset, "_run_state.json")) if req.resume else None
        if resume_state and not resume_state.get("isBatch"):
            eff_req = RunReq(**resume_state["params"])
            run_id = resume_state["runId"]
            resume_phase, resume_units = resume_state["phase"], resume_state["completedUnits"]
        else:
            eff_req, resume_state = req, None
            run_id = req.runId or f"{req.ruleset}-{datetime.datetime.now():%Y%m%d%H%M%S}"
            resume_phase, resume_units = None, None
        title = eff_req.title or run_id
        sleep_days = eff_req.sleepDays if eff_req.sleepDays is not None else int(cfg["sleeping"]["sleepDays"])
        jobs[job_id].update(status="running", runId=run_id, step="start")
        checkpoint_base = {"runId": run_id, "ruleset": eff_req.ruleset, "isBatch": False,
                           "params": eff_req.model_dump()}
        with driver.session() as s:
            result = _run_one(s, job_id, cfg, ruleset=eff_req.ruleset, dataset=eff_req.dataset,
                               user_type_profile=eff_req.userTypeProfile, org_profile=eff_req.orgProfile,
                               sleep_days=sleep_days, min_criticality_rank=eff_req.minCriticalityRank,
                               sod_rules=eff_req.sodRules, query_scope=eff_req.queryScope,
                               query_ids=eff_req.queryIds, evaluate_sod=eff_req.evaluateSod,
                               run_id=run_id, title=title,
                               skip_ruleset_load=eff_req.skipRulesetLoad, skip_materialize=eff_req.skipMaterialize,
                               skip_explain=eff_req.skipExplain, checkpoint_base=checkpoint_base,
                               resume_phase=resume_phase, resume_units=resume_units)
        _clear_checkpoint(_checkpoint_path(eff_req.dataset, "_run_state.json"))
        jobs[job_id].update(status="done", step="done", **result)
    except InterruptedError:
        jobs[job_id].update(status="cancelled", step="cancelled")
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


class RunBatchReq(BaseModel):
    ruleset: str = "kpmg_r3"
    dataset: str
    userTypeProfile: str = "all"
    orgProfiles: list[str] = []      # >=1 Org-Varianten (je Variante ein eigener Lauf) -- leer nur
                                     # erlaubt, wenn resume=true (dann kommen sie aus dem Checkpoint)
    sleepDays: int | None = None
    minCriticalityRank: int = 0
    sodRules: list[str] = []
    queryScope: str = "all"          # s. RunReq.queryScope
    queryIds: list[str] = []         # s. RunReq.queryIds
    evaluateSod: bool = True         # s. RunReq.evaluateSod
    skipRulesetLoad: bool = True
    skipMaterialize: bool = False
    skipExplain: bool = False        # s. RunReq.skipExplain
    resume: bool = False             # s. RunReq.resume


def do_run_batch(job_id: str, req: RunBatchReq):
    try:
        cfg = profiles()
        resume_state = _read_checkpoint(_checkpoint_path(req.dataset, "_run_state.json")) if req.resume else None
        if resume_state and resume_state.get("isBatch"):
            eff_req = RunBatchReq(**resume_state["params"])
            ts = resume_state["ts"]
            results = resume_state.get("completedResults", [])
            start_index = resume_state["variantIndex"]
            resume_phase, resume_units = resume_state["phase"], resume_state["completedUnits"]
        else:
            if not req.orgProfiles:
                raise ValueError("orgProfiles darf nicht leer sein")
            eff_req, resume_state = req, None
            ts = f"{datetime.datetime.now():%Y%m%d%H%M%S}"
            results = []
            start_index = 1
            resume_phase, resume_units = None, None
        sleep_days = eff_req.sleepDays if eff_req.sleepDays is not None else int(cfg["sleeping"]["sleepDays"])
        n = len(eff_req.orgProfiles)
        jobs[job_id].update(status="running", step="start", runs=results)
        with driver.session() as s:
            for i, profile_name in enumerate(eff_req.orgProfiles, start=1):
                if i < start_index:
                    continue    # bereits abgeschlossene Variante -> Ergebnis steckt schon in `results`
                op = next((p for p in cfg["profiles"] if p["name"] == profile_name), None)
                if not op:
                    raise ValueError(f"orgProfile '{profile_name}' unbekannt")
                run_id = f"{eff_req.ruleset}-{ts}-{i}-{_slug(profile_name)}"
                title = _humanize_profile_name(profile_name)
                checkpoint_base = {"runId": run_id, "ruleset": eff_req.ruleset, "isBatch": True,
                                   "params": eff_req.model_dump(), "ts": ts,
                                   "variantIndex": i, "variantTotal": n, "completedResults": results}
                result = _run_one(
                    s, job_id, cfg, ruleset=eff_req.ruleset, dataset=eff_req.dataset,
                    user_type_profile=eff_req.userTypeProfile, org_profile=profile_name,
                    sleep_days=sleep_days, min_criticality_rank=eff_req.minCriticalityRank,
                    sod_rules=eff_req.sodRules, query_scope=eff_req.queryScope,
                    query_ids=eff_req.queryIds, evaluate_sod=eff_req.evaluateSod,
                    run_id=run_id, title=title,
                    # Ruleset nur einmal laden (idempotent, aber unnoetig fuer jede Variante)
                    skip_ruleset_load=eff_req.skipRulesetLoad or i > 1,
                    skip_materialize=eff_req.skipMaterialize, skip_explain=eff_req.skipExplain,
                    step_prefix=f"Variante {i}/{n} ({profile_name}): ",
                    checkpoint_base=checkpoint_base,
                    resume_phase=resume_phase if i == start_index else None,
                    resume_units=resume_units if i == start_index else None,
                )
                results.append(result)
                jobs[job_id]["runs"] = results
        _clear_checkpoint(_checkpoint_path(eff_req.dataset, "_run_state.json"))
        jobs[job_id].update(status="done", step="done", runs=results)
    except InterruptedError:
        jobs[job_id].update(status="cancelled", step="cancelled")
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


class ImportReq(BaseModel):
    dataset: str                    # Ordnername unter data/import/ (= dataset-Id)
    lang: list[str] = []            # Sprach-Schalter (SPRAS/LANGU); leer = Default (IMPORT_LANG)
    skipConvert: bool = False       # .csv liegen schon vor -> Konvertierung ueberspringen
    skipSchema: bool = False        # Migrationen (Constraints/Indizes) ueberspringen
    clearFirst: bool = False        # Dataset im Graph leeren vor Import (Neustart, entfernt Artefakte)
    resume: bool = False            # Ab letztem Checkpoint weitermachen (State-Datei lesen)
    asOf: str | None = None         # 'YYYY-MM-DD' Stichtag (= Downloaddatum der SAP-Extrakte);
                                     # nur bei Erst-Import wirksam (ON CREATE), Default = heute.
                                     # Spaetere Korrektur ueber PUT /datasets/{id}/asof.


def _check_cancel(job_id: str):
    if jobs.get(job_id, {}).get("cancelRequested"):
        raise InterruptedError("Abgebrochen")


# --- Import-Checkpoint (Resume nach Abbruch) ------------------------------------
def _state_path(dataset: str) -> Path:
    return DATA_DIR / dataset / "_import_state.json"

def _read_state(dataset: str) -> dict | None:
    p = _state_path(dataset)
    try:
        return json.loads(p.read_text(encoding="utf-8")) if p.is_file() else None
    except Exception:
        return None

def _write_state(dataset: str, completed: list[str], total: int) -> None:
    state = {"dataset": dataset, "completed": completed, "total": total,
             "updatedAt": datetime.datetime.now().isoformat(timespec="seconds")}
    p = _state_path(dataset)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

def _clear_state(dataset: str) -> None:
    p = _state_path(dataset)
    if p.is_file():
        p.unlink()


def _persist_import_evidence(session, dataset: str, lang: list[str], conv_tables: list[dict]) -> None:
    """Import-Evidenz (Vollstaendigkeitsnachweis, ROADMAP.md) persistieren -- bisher flüchtig
    (Job-Counts) bzw. verworfen (99_validate.cypher lief nur via .consume(), s. run_cypher_path()).
    Ein (:Import)-Knoten je Import-Vorgang (nicht ueberschrieben -> Historie ueber Re-Importe),
    mit Quellzeilen/verworfenen Spalten je Tabelle (aus dem Konverter, convert.py liefert das
    schon) + Graph-Zaehlern je Label/Kantentyp (99_validate.cypher, hier ueber
    run_cypher_path_capturing() erneut ausgefuehrt, um die sonst verworfenen Zeilen zu behalten --
    rein lesend, keine zusaetzlichen Schreib-Kosten). Grundlage fuer Konsistenzcheck I1
    (cypher/checks/import_evidence.cypher). conv_tables leer, wenn skipConvert gesetzt war (dann
    fehlt nur die Quellzeilen-/Spalten-Statistik, die Graph-Zaehler bleiben live und korrekt)."""
    imported_at = datetime.datetime.now()
    key = f"{dataset}|{imported_at.isoformat()}"
    session.run(
        "MERGE (d:Dataset {id:$dataset}) "
        "CREATE (i:Import {key:$key, dataset:$dataset, importedAt:datetime($importedAt), lang:$lang}) "
        "MERGE (d)-[:HAS_IMPORT]->(i)",
        dataset=dataset, key=key, importedAt=imported_at.isoformat(), lang=lang)

    if conv_tables:
        session.run(
            "MATCH (i:Import {key:$key}) UNWIND $tables AS t "
            "CREATE (it:ImportTable {key:$key + '|' + t.table, table:t.table, "
            "  sourceRows:t.rows, droppedColumns:t.dropped}) "
            "MERGE (i)-[:HAS_TABLE]->(it)",
            key=key, tables=[{"table": t["table"], "rows": t["rows"],
                              "dropped": t.get("dropped") or []} for t in conv_tables])

    # DELETED='X' wird bislang nur bei AGR_1251 gefiltert (load/08_authorizations.cypher) -- der
    # Konverter zaehlt das nicht mit (reine CSV-Konvertierung, kein fachlicher Filter), daher
    # separat per LOAD CSV nachgezaehlt. Datei kann fehlen (z. B. skipConvert/Backup&Clear) --
    # dann bleibt filteredRows einfach ungesetzt, kein harter Fehler.
    if any(t["table"] == "agr_1251" for t in conv_tables):
        try:
            n = session.run(
                "LOAD CSV WITH HEADERS FROM $url AS row FIELDTERMINATOR '\t' "
                "WITH row WHERE coalesce(row.DELETED,'') = 'X' RETURN count(*) AS n",
                url=f"file:///{dataset}/agr_1251.csv").single()["n"]
            session.run("MATCH (it:ImportTable {key:$key + '|agr_1251'}) SET it.filteredRows=$n",
                       key=key, n=n)
        except Exception:
            pass

    node_rows, edge_rows = run_cypher_path_capturing(session, LOAD_DIR / "99_validate.cypher",
                                                      {"dataset": dataset})
    session.run(
        "MATCH (i:Import {key:$key}) UNWIND $rows AS r "
        "CREATE (c:ImportNodeCount {key:$key + '|n|' + apoc.text.join(r.labels,','), "
        "  labels:r.labels, count:r.count}) "
        "MERGE (i)-[:HAS_NODE_COUNT]->(c)",
        key=key, rows=[{"labels": r["labels"], "count": r["knoten"]} for r in node_rows])
    session.run(
        "MATCH (i:Import {key:$key}) UNWIND $rows AS r "
        "CREATE (c:ImportEdgeCount {key:$key + '|e|' + r.type, type:r.type, count:r.count}) "
        "MERGE (i)-[:HAS_EDGE_COUNT]->(c)",
        key=key, rows=[{"type": r["kante"], "count": r["anzahl"]} for r in edge_rows])


def do_import(job_id: str, req: ImportReq):
    try:
        folder = DATA_DIR / req.dataset
        if not folder.is_dir():
            raise ValueError(f"Import-Ordner fehlt: data/import/{req.dataset}")
        lang = req.lang or DEFAULT_LANG
        as_of = (datetime.date.fromisoformat(req.asOf) if req.asOf
                 else (_infer_dataset_asof(req.dataset) or datetime.date.today()))
        jobs[job_id].update(status="running", dataset=req.dataset, step="start")

        # 0. Optional: Dataset vor Import leeren (Artefakte aus früheren Läufen entfernen)
        if req.clearFirst:
            jobs[job_id]["step"] = "clear"
            with driver.session() as s:
                run_file(s, "admin/clear_dataset.cypher", {"dataset": req.dataset})
            _clear_state(req.dataset)

        # Resume: bereits erledigte Load-Steps aus State-Datei lesen
        completed_steps: set[str] = set()
        completed_list: list[str] = []
        if req.resume and not req.clearFirst:
            state = _read_state(req.dataset)
            if state:
                completed_list = state.get("completed", [])
                completed_steps = set(completed_list)
                jobs[job_id]["resumedFrom"] = completed_list[-1] if completed_list else ""

        # 1. Konvertieren (SE16-.txt -> .csv; Minimalset-Pruefung + Credential-Denylist im Konverter)
        if not req.skipConvert:
            jobs[job_id].update(step="convert", convertFile="", convertDone=0, convertTotal=0)

            def _conv_progress(done, total, stem):
                jobs[job_id].update(convertDone=done, convertTotal=total, convertFile=stem)

            conv = convert.convert_folder(folder, required_config=CONFIG_DIR / "required_tables.json",
                                          on_progress=_conv_progress)
            jobs[job_id]["converted"] = conv["converted"]
            jobs[job_id]["skipped"] = conv.get("skipped", 0)
            jobs[job_id]["missingOptional"] = conv["missingOptional"]

        _check_cancel(job_id)
        with driver.session() as s:
            # 2. Schema sicherstellen (idempotente CREATE ... IF NOT EXISTS)
            if not req.skipSchema:
                jobs[job_id]["step"] = "schema"
                for f in sorted(MIGRATIONS_DIR.glob("*.cypher")):
                    run_cypher_path(s, f, {})
            _check_cancel(job_id)
            # 3. Laden (Reihenfolge = Dateiname), mit dataset + lang + asOf (nur 00_dataset.cypher
            # nutzt asOf, ungenutzte gebundene Parameter in den anderen Loadern sind unschaedlich)
            load_files = sorted(LOAD_DIR.glob("*.cypher"))
            for idx, f in enumerate(load_files, 1):
                jobs[job_id].update(step=f"load {f.name}", stepNum=idx, stepTotal=len(load_files))
                if f.name in completed_steps:
                    jobs[job_id]["stepSkipped"] = True
                    continue                              # bereits erledigt -> ueberspringen
                jobs[job_id].pop("stepSkipped", None)
                _check_cancel(job_id)
                try:
                    counters = run_cypher_path(s, f, {"dataset": req.dataset, "lang": lang, "asOf": as_of})
                except Exception as e:
                    if "NoSuchFileException" in str(e):
                        # Optionale Tabelle fehlt im Export -> Schritt überspringen, kein Abbruch
                        jobs[job_id]["lastWarning"] = f"Datei fehlt, Schritt übersprungen: {f.name}"
                        counters = {"nodes_created": 0, "relationships_created": 0, "properties_set": 0}
                    else:
                        raise
                jobs[job_id]["lastNodes"] = counters["nodes_created"]
                jobs[job_id]["lastRels"] = counters["relationships_created"]
                completed_list.append(f.name)
                _write_state(req.dataset, completed_list, len(load_files))
            jobs[job_id].pop("stepNum", None)
            jobs[job_id].pop("stepTotal", None)
            # 4. Validieren (eigene Zaehler statt Konsolen-Output) + Import-Evidenz persistieren
            jobs[job_id]["step"] = "validate"
            rec = s.run(
                "OPTIONAL MATCH (u:User {dataset:$d}) WITH count(u) AS users "
                "OPTIONAL MATCH (r:Role {dataset:$d}) WITH users, count(r) AS roles "
                "OPTIONAL MATCH (a:Authorization {dataset:$d}) RETURN users, roles, count(a) AS auths",
                d=req.dataset).single()
            _persist_import_evidence(s, req.dataset, lang,
                                     conv.get("tables", []) if not req.skipConvert else [])
        _clear_state(req.dataset)   # vollstaendig -> State-Datei nicht mehr benoetigt
        jobs[job_id].update(status="done", step="done",
                            users=rec["users"], roles=rec["roles"], auths=rec["auths"])
    except InterruptedError:
        jobs[job_id].update(status="cancelled", step="cancelled")
        # State-Datei bleibt erhalten -> Resume moeglich
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


def do_explain(job_id: str, run_id: str):
    """Evidenz (VIA_ROLE/VIA_PROFILE + conflictType) fuer einen bestehenden Lauf nachrechnen.
    Scope (ruleset/dataset/asOf) kommt aus dem (:Run)-Knoten. Fortschritt/Resume ueber
    _explain_one() wie die explain-Phase eines Laufs -- eigene, nach runId benannte
    Checkpoint-Datei (statt der dataset-weiten _run_state.json), damit ein parallel laufender
    "echter" Lauf desselben Datasets nicht denselben Checkpoint ueberschreibt. Ein zuvor
    abgebrochener Aufruf fuer denselben Lauf wird automatisch fortgesetzt (kein Formular/keine
    Nutzer-Entscheidung noetig -- der einzige Parameter ist der bestehende Lauf selbst)."""
    try:
        with driver.session() as s:
            run = s.run("MATCH (r:Run {runId:$id}) RETURN r.ruleset AS ruleset, r.dataset AS dataset, "
                        "r.asOf AS asOf", id=run_id).single()
            if not run:
                raise ValueError(f"Lauf '{run_id}' nicht gefunden")
            jobs[job_id].update(status="running", step="explain", runId=run_id)
            state_path = _checkpoint_path(run["dataset"], f"_explain_state_{run_id}.json")
            resume_state = _read_checkpoint(state_path)
            resume_units = set(resume_state["completedUnits"]) if resume_state else set()
            _explain_one(s, job_id, ruleset=run["ruleset"], dataset=run["dataset"], as_of=run["asOf"],
                        run_id=run_id, state_path=state_path,
                        checkpoint_extra={"runId": run_id, "ruleset": run["ruleset"]},
                        resume_units=resume_units)
            _clear_checkpoint(state_path)
            rec = s.run("MATCH (f:SoDConflict {runId:$id}) RETURN count(f) AS findings, "
                        "sum(CASE WHEN f.conflictType='intra' THEN 1 ELSE 0 END) AS intra, "
                        "sum(CASE WHEN f.conflictType='inter' THEN 1 ELSE 0 END) AS inter", id=run_id).single()
        jobs[job_id].update(status="done", step="done",
                            findings=rec["findings"], intra=rec["intra"], inter=rec["inter"])
    except InterruptedError:
        jobs[job_id].update(status="cancelled", step="cancelled")
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


def do_delete_run(job_id: str, run_id: str):
    """Loescht einen einzelnen Auswertungslauf (Run + Findings); Dataset bleibt unberuehrt."""
    try:
        jobs[job_id].update(status="running", step="delete", runId=run_id)
        with driver.session() as s:
            run_file(s, "admin/delete_run.cypher", {"runId": run_id})
        jobs[job_id].update(status="done", step="done")
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


def do_clear(job_id: str, dataset: str):
    try:
        jobs[job_id].update(status="running", step="clear", dataset=dataset)
        with driver.session() as s:
            run_file(s, "admin/clear_dataset.cypher", {"dataset": dataset})
            rec = s.run("MATCH (n {dataset:$d}) RETURN count(n) AS remaining", d=dataset).single()
        _clear_state(dataset)
        jobs[job_id].update(status="done", step="done", remaining=rec["remaining"])
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


def do_reset(job_id: str):
    try:
        jobs[job_id].update(status="running", step="reset")
        with driver.session() as s:
            run_file(s, "admin/reset_data.cypher", {})
            rec = s.run("MATCH (n) WHERE NOT n:Query AND NOT n:SoDRule AND NOT n:AuthReq "
                        "AND NOT n:Clause AND NOT n:__Neo4jMigration "
                        "RETURN count(n) AS remaining").single()
            q = s.run("MATCH (q:Query) RETURN count(q) AS queries").single()
        jobs[job_id].update(status="done", step="done",
                            remaining=rec["remaining"], rulesetQueries=q["queries"])
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


# --- Backup/Restore (Quelldaten-Ebene) -------------------------------------------------
# Backup = ZIP der konvertierten, credential-bereinigten .csv eines Datasets (+ manifest.json);
# Restore = entpacken + deterministischer Re-Import. Online, transportabel (eine Datei), trust-aware
# (nur bereinigte .csv, nie die rohen .txt). Findings sind regenerierbar (neu auswerten nach Restore).
_SAFE_NAME = re.compile(r"^[A-Za-z0-9._-]+$")


def _backup_path(file: str) -> Path:
    if not _SAFE_NAME.match(file) or not file.endswith(".zip"):
        raise HTTPException(400, "ungueltiger Dateiname")
    p = (BACKUP_DIR / file).resolve()
    if p.parent != BACKUP_DIR.resolve() or not p.is_file():
        raise HTTPException(404, "Backup nicht gefunden")
    return p


def do_backup(job_id: str, dataset: str, clear: bool):
    try:
        folder = DATA_DIR / dataset
        if not folder.is_dir():
            raise ValueError(f"Import-Ordner fehlt: data/import/{dataset}")
        jobs[job_id].update(status="running", step="backup", dataset=dataset)
        csvs = sorted(folder.glob("*.csv"))
        if not csvs:   # nur .txt vorhanden -> erst konvertieren (bereinigte .csv erzeugen)
            convert.convert_folder(folder, required_config=CONFIG_DIR / "required_tables.json")
            csvs = sorted(folder.glob("*.csv"))
        if not csvs:
            raise ValueError("keine .csv/.txt zum Sichern gefunden")

        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        name = f"{dataset}__{ts}.zip"
        manifest = {"dataset": dataset, "createdAt": datetime.datetime.now().isoformat(timespec="seconds"),
                    "tables": [{"table": c.stem, "bytes": c.stat().st_size} for c in csvs]}
        with zipfile.ZipFile(BACKUP_DIR / name, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
            for c in csvs:
                z.write(c, arcname=f"csv/{c.name}")
        jobs[job_id].update(backup=name, sizeBytes=(BACKUP_DIR / name).stat().st_size, tables=len(csvs))

        if clear:   # "Backup & Clear": Graph des Datasets leeren (Ruleset/Schema bleiben)
            jobs[job_id]["step"] = "clear"
            with driver.session() as s:
                run_file(s, "admin/clear_dataset.cypher", {"dataset": dataset})
            jobs[job_id]["cleared"] = True
        jobs[job_id].update(status="done", step="done")
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


def do_restore(job_id: str, path: Path, dataset: str | None):
    try:
        jobs[job_id].update(status="running", step="unpack")
        with zipfile.ZipFile(path) as z:
            manifest = json.loads(z.read("manifest.json"))
            target = dataset or manifest["dataset"]
            if not _SAFE_NAME.match(target):
                raise ValueError(f"ungueltiger Ziel-Dataset-Name: {target}")
            dest = DATA_DIR / target
            dest.mkdir(parents=True, exist_ok=True)
            for n in z.namelist():
                if n.startswith("csv/") and n.endswith(".csv"):
                    (dest / Path(n).name).write_bytes(z.read(n))
        # .csv liegen jetzt vor -> Re-Import ohne Konvertierung (Schema idempotent sicherstellen)
        do_import(job_id, ImportReq(dataset=target, skipConvert=True, skipSchema=False))
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


# --- Backup/Restore (Auswertungslauf-Ebene) --------------------------------------------
# Lauf-Backup = ZIP aus manifest.json (Provenienz inkl. Dataset-uid) + run.json + findings.json.
# Getrennt von Quelldaten-Backups (oben): sichert das ABGELEITETE Ergebnis eines Laufs, kein
# Re-Import. Restore prueft die Dataset-uid aus dem Manifest gegen die aktuelle Dataset-uid —
# Mismatch bedeutet "das Dataset wurde seit dem Lauf neu befuellt", Restore dann nur mit
# force=true (Warnung wird vom Frontend angezeigt).
def _run_backup_path(file: str) -> Path:
    if not _SAFE_NAME.match(file) or not file.endswith(".zip"):
        raise HTTPException(400, "ungueltiger Dateiname")
    p = (RUN_BACKUP_DIR / file).resolve()
    if p.parent != RUN_BACKUP_DIR.resolve() or not p.is_file():
        raise HTTPException(404, "Lauf-Backup nicht gefunden")
    return p


def _dataset_uid(s, dataset: str) -> str | None:
    """Aktuelle Dataset-uid; legt sie nach (lazy backfill), falls das Dataset noch keine hat."""
    rec = s.run("MATCH (d:Dataset {id:$id}) SET d.uid = coalesce(d.uid, randomUUID()) "
                "RETURN d.uid AS uid", id=dataset).single()
    return rec["uid"] if rec else None


def _infer_dataset_asof(dataset: str) -> datetime.date | None:
    """Leitet den Stichtag (= Downloaddatum der SAP-Extrakte) aus den Dateizeitstempeln des
    Import-Ordners ab -- bevorzugt die rohen SE16-.txt-Exporte (alle Tabellen eines Extrakts
    teilen sich praktisch immer denselben Exporttag, da sie in einer Sitzung gezogen werden),
    sonst ersatzweise die konvertierten .csv. None, wenn der Ordner fehlt/leer ist (z. B. der
    Quellordner wurde nach dem Import geloescht) -- dann faellt der Aufrufer auf 'heute' zurueck.
    Nimmt bewusst den FRUEHESTEN Zeitstempel (nicht den spaetesten): einzelne Tabellen koennen
    minimal nacheinander exportiert/konvertiert worden sein, der frueheste liegt am naechsten am
    eigentlichen Download-Zeitpunkt."""
    folder = DATA_DIR / dataset
    if not folder.is_dir():
        return None
    files = list(folder.glob("*.txt")) or list(folder.glob("*.csv"))
    if not files:
        return None
    return datetime.date.fromtimestamp(min(f.stat().st_mtime for f in files))


def _dataset_asof(s, dataset: str):
    """Stichtag des Datasets (= Downloaddatum der SAP-Extrakte, Neo4j-date) -- KEIN Lauf-/Check-
    Parameter mehr, sondern eine Eigenschaft des Datasets selbst (s. load/00_dataset.cypher).
    Legt ihn nach (lazy backfill), falls ein aelterer Import noch keinen hat: bevorzugt aus den
    Dateizeitstempeln des Import-Ordners abgeleitet (_infer_dataset_asof), nur wenn der Ordner
    nicht mehr existiert als letzter Ausweg 'heute'. Aendern danach nur bewusst ueber PUT
    /datasets/{id}/asof (globale Variable), nicht je Lauf/Check."""
    rec = s.run("MATCH (d:Dataset {id:$id}) RETURN d.asOf AS asOf", id=dataset).single()
    if not rec:
        raise HTTPException(404, f"Dataset '{dataset}' nicht gefunden")
    if rec["asOf"] is not None:
        return rec["asOf"]
    inferred = _infer_dataset_asof(dataset) or datetime.date.today()
    rec2 = s.run("MATCH (d:Dataset {id:$id}) SET d.asOf = coalesce(d.asOf, $asOf) "
                 "RETURN d.asOf AS asOf", id=dataset, asOf=inferred).single()
    return rec2["asOf"]


def do_backup_run(job_id: str, run_id: str):
    try:
        jobs[job_id].update(status="running", step="backup", runId=run_id)
        with driver.session() as s:
            rec = s.run("MATCH (r:Run {runId:$id}) RETURN r AS run", id=run_id).single()
            if not rec:
                raise ValueError(f"Lauf '{run_id}' nicht gefunden")
            run_d = jsonable(dict(rec["run"]))
            dataset_uid = _dataset_uid(s, run_d["dataset"])
            findings = [jsonable(dict(r)) for r in s.run(
                "MATCH (u:User)-[:VIOLATES]->(f:SoDConflict {runId:$id})-[:BASED_ON]->(rule:SoDRule) "
                "RETURN u.id AS user, f.ruleId AS rule, f.criticality AS criticality, "
                "f.criticalityRank AS criticalityRank, f.asOf AS asOf, f.reasonCode AS reasonCode, "
                "coalesce(f.userSleeping,false) AS sleeping, "
                "coalesce(f.lastLogonKnown,true) AS lastLogonKnown, f.conflictType AS conflictType",
                id=run_id)]
        RUN_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        name = f"{run_id}__{ts}.run.zip"
        manifest = {"runId": run_id, "dataset": run_d["dataset"], "datasetUid": dataset_uid,
                    "ruleset": run_d.get("ruleset"), "title": run_d.get("title"), "asOf": run_d.get("asOf"),
                    "generatedAt": run_d.get("generatedAt"),
                    "createdAt": datetime.datetime.now().isoformat(timespec="seconds"),
                    "findingsCount": len(findings)}
        with zipfile.ZipFile(RUN_BACKUP_DIR / name, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
            z.writestr("run.json", json.dumps(run_d, ensure_ascii=False, indent=2))
            z.writestr("findings.json", json.dumps(findings, ensure_ascii=False, indent=2))
        jobs[job_id].update(status="done", step="done", backup=name, findings=len(findings),
                            sizeBytes=(RUN_BACKUP_DIR / name).stat().st_size)
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


def do_restore_run(job_id: str, path: Path, force: bool):
    try:
        jobs[job_id].update(status="running", step="check")
        with zipfile.ZipFile(path) as z:
            manifest = json.loads(z.read("manifest.json"))
            run_d = json.loads(z.read("run.json"))
            findings = json.loads(z.read("findings.json"))
        with driver.session() as s:
            current_uid = _dataset_uid(s, manifest["dataset"])
            if not force and current_uid != manifest.get("datasetUid"):
                raise ValueError(
                    f"Dataset '{manifest['dataset']}' hat sich seit diesem Lauf-Backup geaendert "
                    "(anderer Datenstand) — Wiederherstellung nur mit ausdruecklicher Bestaetigung.")
            jobs[job_id]["step"] = "restore"
            run_file(s, "admin/restore_run.cypher",
                     {"run": run_d, "findings": findings, "runId": manifest["runId"]})
            rec = s.run("MATCH (f:SoDConflict {runId:$id}) RETURN count(f) AS restored",
                        id=manifest["runId"]).single()
        jobs[job_id].update(status="done", step="done", runId=manifest["runId"],
                            restored=rec["restored"], total=len(findings))
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))


def do_upload_import(job_id: str, zip_path: str, dataset: str, lang: list[str], skip_schema: bool,
                      as_of: str | None = None, clear_first: bool = False):
    """Entpackt ein hochgeladenes ZIP (.csv und/oder .txt) nach data/import/<dataset> und importiert.
    .txt vorhanden -> konvertieren; nur .csv -> Konvertierung ueberspringen."""
    try:
        jobs[job_id].update(status="running", step="unpack", dataset=dataset)
        dest = DATA_DIR / dataset
        dest.mkdir(parents=True, exist_ok=True)
        extracted, has_txt = 0, False
        with zipfile.ZipFile(zip_path) as z:
            for n in z.namelist():
                low = n.lower()
                if low.endswith(".csv") or low.endswith(".txt"):
                    (dest / Path(n).name).write_bytes(z.read(n))
                    extracted += 1
                    has_txt = has_txt or low.endswith(".txt")
        if not extracted:
            raise ValueError("ZIP enthielt keine .csv/.txt-Dateien")
        do_import(job_id, ImportReq(dataset=dataset, lang=lang,
                                    skipConvert=not has_txt, skipSchema=skip_schema, asOf=as_of,
                                    clearFirst=clear_first))
    except Exception as e:  # noqa: BLE001
        jobs[job_id].update(status="error", error=str(e))
        _log_job_error(job_id, str(e))
    finally:
        try:
            os.remove(zip_path)
        except OSError:
            pass


@app.get("/health")
def health():
    with driver.session() as s:
        s.run("RETURN 1").consume()
    return {"status": "ok"}


@app.get("/datasets")
def datasets():
    with driver.session() as s:
        return [r["id"] for r in s.run("MATCH (d:Dataset) RETURN d.id AS id ORDER BY id")]


class DatasetAsOfReq(BaseModel):
    asOf: str   # 'YYYY-MM-DD'


@app.get("/datasets/{datasetId}/asof")
def dataset_asof(datasetId: str):
    """Stichtag des Datasets (= Downloaddatum der SAP-Extrakte) -- globale Eigenschaft, kein
    Lauf-/Check-Filter mehr (s. _dataset_asof()). UI zeigt diesen Wert nur an/erlaubt ihn
    gezielt zu korrigieren (PUT), nicht je Auswertung neu zu waehlen."""
    with driver.session() as s:
        return {"dataset": datasetId, "asOf": jsonable(_dataset_asof(s, datasetId))}


@app.put("/datasets/{datasetId}/asof")
def set_dataset_asof(datasetId: str, req: DatasetAsOfReq):
    """Bewusste Korrektur des Dataset-Stichtags (globale Variable) -- z. B. falls beim Import
    ein falsches Downloaddatum gesetzt wurde. Wirkt auf ALLE folgenden Laeufe/Checks dieses
    Datasets, nicht nur auf den naechsten."""
    as_of = datetime.date.fromisoformat(req.asOf)
    with driver.session() as s:
        rec = s.run("MATCH (d:Dataset {id:$id}) SET d.asOf = $asOf RETURN d.asOf AS asOf",
                    id=datasetId, asOf=as_of).single()
        if not rec:
            raise HTTPException(404, f"Dataset '{datasetId}' nicht gefunden")
    return {"dataset": datasetId, "asOf": jsonable(rec["asOf"])}


@app.get("/datasets/{datasetId}/import-files")
def import_files_info(datasetId: str):
    """Groesse der rohen Quelldateien (.txt + .csv) im Import-Ordner des Datasets."""
    folder = DATA_DIR / datasetId
    if not folder.is_dir():
        return {"dataset": datasetId, "files": 0, "bytes": 0}
    files = list(folder.glob("*.txt")) + list(folder.glob("*.csv"))
    return {"dataset": datasetId, "files": len(files),
            "bytes": sum(f.stat().st_size for f in files)}


@app.delete("/datasets/{datasetId}/import-files")
def delete_import_files(datasetId: str):
    """Loescht den kompletten data/import/<dataset>/-Ordner (nach erfolgtem Backup nicht
    mehr benoetigt; fuer Re-Import reicht das Backup-ZIP per Upload)."""
    if not _SAFE_NAME.match(datasetId):
        raise HTTPException(400, "ungueltiger Dataset-Name")
    folder = DATA_DIR / datasetId
    if not folder.is_dir():
        raise HTTPException(404, f"Import-Ordner '{datasetId}' nicht gefunden")
    files = list(folder.glob("*.txt")) + list(folder.glob("*.csv"))
    freed = sum(f.stat().st_size for f in files)
    deleted = len(files)
    shutil.rmtree(folder)
    return {"dataset": datasetId, "deleted": deleted, "freedBytes": freed}


@app.get("/users/{userId}")
def user_detail(userId: str, runId: str):
    """Kurzprofil eines Users (Name/Typ/Status) fuer die Kontext-Chips bei nutzerzentrischer
    Auswahl — dataset wird ueber den Lauf aufgeloest."""
    with driver.session() as s:
        run = s.run("MATCH (r:Run {runId:$id}) RETURN r.dataset AS dataset", id=runId).single()
        if not run:
            raise HTTPException(404, f"Lauf '{runId}' nicht gefunden")
        rec = s.run(
            "MATCH (u:User {id:$uid, dataset:$dataset}) "
            "RETURN u.id AS id, coalesce(u.name,'') AS name, "
            "  CASE WHEN 'Dialog' IN labels(u) THEN 'Dialog' WHEN 'System' IN labels(u) THEN 'System' "
            "       WHEN 'Service' IN labels(u) THEN 'Service' WHEN 'Communication' IN labels(u) THEN 'Communication' "
            "       WHEN 'Reference' IN labels(u) THEN 'Reference' ELSE '?' END AS typ, "
            "  CASE WHEN 'Locked' IN labels(u) THEN 'gesperrt' ELSE 'aktiv' END AS status",
            uid=userId, dataset=run["dataset"]).single()
        if not rec:
            raise HTTPException(404, f"User '{userId}' nicht gefunden")
        return dict(rec)


@app.get("/roles/{roleId}")
def role_detail(roleId: str, runId: str, user: str | None = None):
    """Rollen-Detailseite: Stammdaten + TCodes (Menue + effektive S_TCODE) + Berechtigungsobjekte.
    dataset ueber den Lauf aufgeloest; optional ?user= fuer die Gueltigkeit der Zuweisung dieses
    Users (Rolle wird i. d. R. aus dem Root-Cause eines konkreten Users geoeffnet)."""
    with driver.session() as s:
        run = s.run("MATCH (r:Run {runId:$id}) RETURN r.dataset AS dataset", id=runId).single()
        if not run:
            raise HTTPException(404, f"Lauf '{runId}' nicht gefunden")
        ds = run["dataset"]
        rec = s.run(
            "MATCH (r:Role {id:$rid, dataset:$ds}) "
            "OPTIONAL MATCH (r)-[:HAS_PROFILE]->(p:Profile) "
            "WITH r, collect(DISTINCT p.id) AS profiles "
            "OPTIONAL MATCH (u:User)-[:ASSIGNED_TO]->(r) "
            "WITH r, profiles, count(DISTINCT u) AS userCount "
            # Ersteller/Aenderer sind SAP-User-Kuerzel (AGR_DEFINE.CREATE_USR/CHANGE_USR) --
            # koennen, muessen aber nicht als :User im Dataset vorhanden sein (z. B. Basis-Team
            # ohne Dialog-Zugang im Extrakt). Name optional aus V_USERNAME (User.name).
            "OPTIONAL MATCH (cu:User {id:r.createUsr, dataset:$ds}) "
            "OPTIONAL MATCH (chu:User {id:r.changeUsr, dataset:$ds}) "
            "RETURN r.id AS id, coalesce(r.text,'') AS text, r.parentAgr AS parentAgr, "
            "  r.profileGenerated AS profileGenerated, r.profileState AS profileState, "
            "  r.createUsr AS createUsr, coalesce(cu.name,'') AS createUsrName, r.createDat AS createDat, "
            "  r.changeUsr AS changeUsr, coalesce(chu.name,'') AS changeUsrName, r.changeDat AS changeDat, "
            "  ('Composite' IN labels(r)) AS composite, "
            "  (r.parentAgr IS NOT NULL) AS derived, "
            "  profiles, userCount",
            rid=roleId, ds=ds).single()
        if not rec:
            raise HTTPException(404, f"Rolle '{roleId}' nicht gefunden")
        out = jsonable(dict(rec))
        if user:
            v = s.run(
                "MATCH (u:User {id:$u, dataset:$ds})-[g:ASSIGNED_TO]->(r:Role {id:$rid, dataset:$ds}) "
                "RETURN g.validFrom AS validFrom, g.validTo AS validTo, coalesce(u.name,'') AS userName LIMIT 1",
                u=user, ds=ds, rid=roleId).single()
            out["userValidFrom"] = jsonable(v["validFrom"]) if v else None
            out["userValidTo"] = jsonable(v["validTo"]) if v else None
            out["forUser"] = user
            out["forUserName"] = v["userName"] if v else ""
        out["menuTcodes"] = [r["t"] for r in s.run(
            "MATCH (r:Role {id:$rid, dataset:$ds})-[:HAS_MENU]->(t:Transaction) "
            "RETURN DISTINCT t.id AS t ORDER BY t", rid=roleId, ds=ds)]
        out["authTcodes"] = [r["tc"] for r in s.run(
            "MATCH (r:Role {id:$rid, dataset:$ds})-[:HAS_AUTH]->(a:Authorization {object:'S_TCODE'}) "
            "WITH apoc.any.property(a,'f_TCD') AS tcds WHERE tcds IS NOT NULL "
            "UNWIND tcds AS tc RETURN DISTINCT tc ORDER BY tc", rid=roleId, ds=ds)]
        out["objects"] = [dict(r) for r in s.run(
            "MATCH (r:Role {id:$rid, dataset:$ds})-[:HAS_AUTH]->(a:Authorization)-[:FOR_OBJECT]->(o:AuthObject) "
            "RETURN o.id AS object, coalesce(o.text,'') AS text, count(DISTINCT a) AS instances "
            "ORDER BY object", rid=roleId, ds=ds)]
        return out


# Gemeinsame Spalten fuer jede anklickbare "N User"-Kennzahl (Rollen-Detailseite, Konsistenzcheck-
# Drilldown): ID/Name/Typ/Benutzergruppe/Letzter Login/Sleeping -- dieselbe Sleeping-Definition
# wie im SoD-Root-Cause (s. evaluate_sod_one.cypher: lastLogonKnown/userSleeping).
_USER_ENRICH_RETURN = (
    "RETURN u.id AS id, coalesce(u.name,'') AS name, "
    "  CASE WHEN 'Dialog' IN labels(u) THEN 'Dialog' WHEN 'System' IN labels(u) THEN 'System' "
    "       WHEN 'Service' IN labels(u) THEN 'Service' WHEN 'Communication' IN labels(u) THEN 'Communication' "
    "       WHEN 'Reference' IN labels(u) THEN 'Reference' ELSE '?' END AS userType, "
    "  coalesce(u.userGroup,'') AS userGroup, u.lastLogon AS lastLogon, "
    "  (u.lastLogon IS NOT NULL) AS lastLogonKnown, "
    "  (u.lastLogon IS NOT NULL AND u.lastLogon < ($asOf - duration({days:$sleepDays}))) AS sleeping "
    "ORDER BY id"
)


@app.get("/roles/{roleId}/users")
def role_users(roleId: str, runId: str):
    """Liste der einem Role zugewiesenen User fuer den anklickbaren "Zugewiesene User"-Zaehler
    der Rollen-Detailseite -- dataset ueber den Lauf aufgeloest, Spalten wie POST /users/list."""
    with driver.session() as s:
        run = s.run("MATCH (r:Run {runId:$id}) RETURN r.dataset AS dataset", id=runId).single()
        if not run:
            raise HTTPException(404, f"Lauf '{runId}' nicht gefunden")
        ds = run["dataset"]
        as_of = _dataset_asof(s, ds)
        sleep_days = profiles()["sleeping"]["sleepDays"]
        return [jsonable(dict(r)) for r in s.run(
            "MATCH (u:User {dataset:$ds})-[:ASSIGNED_TO]->(:Role {id:$rid, dataset:$ds}) " + _USER_ENRICH_RETURN,
            rid=roleId, ds=ds, asOf=as_of, sleepDays=sleep_days)]


class UsersListReq(BaseModel):
    dataset: str
    ids: list[str]


@app.post("/users/list")
def users_list(req: UsersListReq):
    """Generisches Nutzer-Enrichment fuer eine gegebene ID-Liste -- speist z. B. den Klick auf eine
    "N User"-Kennzahl im Konsistenzcheck-Ergebnis (die IDs kommen dort aus der bereits geladenen
    Detailtabelle, dataset direkt vom Client). Gleiche Spalten wie GET /roles/{id}/users."""
    with driver.session() as s:
        as_of = _dataset_asof(s, req.dataset)
        sleep_days = profiles()["sleeping"]["sleepDays"]
        return [jsonable(dict(r)) for r in s.run(
            "MATCH (u:User {dataset:$ds}) WHERE u.id IN $ids " + _USER_ENRICH_RETURN,
            ds=req.dataset, ids=req.ids, asOf=as_of, sleepDays=sleep_days)]


@app.get("/roles/{roleId}/can-do")
def role_can_do(roleId: str, runId: str):
    """Rollenzentrisch (lauf-unabhaengig): welche Einzelberechtigungs-Queries erfuellt die Rolle
    allein, und welche SoD-Regeln kann sie allein ausloesen (Intra-Rollen-Konflikt). Ruleset aus
    dem Lauf. Kann etwas dauern (eine Rolle x alle Ruleset-Queries) -> Frontend laedt lazy."""
    with driver.session() as s:
        run = s.run("MATCH (r:Run {runId:$id}) RETURN r.ruleset AS ruleset, r.dataset AS dataset",
                    id=runId).single()
        if not run:
            raise HTTPException(404, f"Lauf '{runId}' nicht gefunden")
        q_stmt = split_statements((CYPHER_DIR / "roles/role_can_do.cypher").read_text(encoding="utf-8"))[0]
        queries = [dict(r) for r in s.run(q_stmt, ruleset=run["ruleset"], dataset=run["dataset"], roleId=roleId)]
        provided = [q["id"] for q in queries]
        r_stmt = split_statements((CYPHER_DIR / "roles/role_sod_rules.cypher").read_text(encoding="utf-8"))[0]
        rules = [dict(r) for r in s.run(r_stmt, ruleset=run["ruleset"], providedIds=provided)]
        return {"role": roleId, "queries": queries, "sodRules": rules}


@app.get("/profiles")
def profiles_meta():
    """Speist die Formular-Dropdowns der UI (datengetrieben aus config + rules)."""
    cfg = profiles()
    return {
        "rulesets": list_rulesets(),
        "defaultRuleset": cfg.get("defaults", {}).get("ruleset"),
        "orgProfiles": [{"name": p["name"], "description": p.get("description", "")} for p in cfg["profiles"]],
        "userTypeProfiles": [{"name": p["name"], "description": p.get("description", "")} for p in cfg["userTypeProfiles"]],
        "scopeProfiles": [{"name": p["name"], "description": p.get("description", "")} for p in cfg.get("scopeProfiles", [])],
        "sleepDays": cfg["sleeping"]["sleepDays"],
    }


# --- Org-Varianten-Editor (Admin) -------------------------------------------------------
# Schreibt NIE in die Vendor-Datei (analysis_profiles.json) — neue/bearbeitete Org-Profile landen
# in einem globalen Overlay (analysis_profiles.custom.json), das profiles() zusaetzlich einliest
# (s. oben). "standard"/"uebergreifend" sind die zwei garantierten Basis-Varianten und bewusst
# geschuetzt (nicht editierbar/loeschbar) — PROTECTED_ORG_PROFILES.
class OrgCriterionReq(BaseModel):
    field: str
    op: str                      # AND | OR | RANGE
    values: list[str] | None = None
    rangeFrom: str | None = None
    rangeTo: str | None = None


class OrgProfileEditReq(BaseModel):
    description: str | None = None
    criteria: list[OrgCriterionReq] = []
    newName: str | None = None       # optionaler neuer Name -> Variante umbenennen (nur eigene)


class OrgProfileCreateReq(OrgProfileEditReq):
    name: str


def _org_filters_from_criteria(criteria: list[OrgCriterionReq]) -> dict:
    filters: dict = {}
    for c in criteria:
        if c.op == "RANGE":
            filters[c.field] = {"op": "RANGE", "from": c.rangeFrom, "to": c.rangeTo}
        else:
            filters[c.field] = {"op": c.op, "values": c.values or []}
    return filters


@app.get("/admin/org-profiles")
def admin_list_org_profiles():
    """Org-Profile (Vendor + Overlay) fuer die Admin-Seite 'Org-Varianten'; 'protected' markiert
    die zwei garantierten Basis-Varianten (nicht editierbar/loeschbar)."""
    cfg = profiles()
    custom_names = {p["name"] for p in _load_json_list(ORG_PROFILES_CUSTOM_PATH)}
    return [{"name": p["name"], "description": p.get("description", ""), "org": p["org"],
             "protected": p["name"] in PROTECTED_ORG_PROFILES,
             "source": "custom" if p["name"] in custom_names else "vendor"}
            for p in cfg["profiles"]]


@app.get("/admin/org-profiles/org-fields")
def admin_org_fields(dataset: str):
    """Org-Felder eines Datasets, auf die sich ein Kriterium sinnvoll einschraenken laesst —
    speist die Feld-Auswahl beim Anlegen eines Org-Kriteriums. Die OrgField-Registry (aus USORG)
    enthaelt alle 50+ moeglichen Org-Ebenen, viele davon kommen im konkreten Berechtigungskonzept
    aber gar nicht oder nur mit echtem '*' (unbeschraenkt) vor -- ein Kriterium darauf waere nie
    waehlbar (org-field-values liefert dann leer). Deshalb hier vorab filtern auf Felder, bei
    denen mindestens eine Authorization einen konkreten (nicht-'*') Wert traegt."""
    with driver.session() as s:
        rows = s.run(
            "MATCH (of:OrgField {dataset:$d}) "
            "WHERE EXISTS { MATCH (a:Authorization {dataset:$d}) "
            "  WHERE apoc.any.property(a,'f_'+of.field) IS NOT NULL "
            "    AND any(v IN apoc.any.property(a,'f_'+of.field) WHERE v <> '*') } "
            "RETURN of.field AS field ORDER BY of.field", d=dataset)
        return [r["field"] for r in rows]


@app.get("/admin/org-profiles/org-field-values")
def admin_org_field_values(dataset: str, field: str):
    """Tatsaechlich im Dataset vorkommende Werte eines Org-Felds (aus den Authorization-
    Feldwerten) — speist die Werte-Auswahl beim Anlegen eines Org-Kriteriums, statt Freitext."""
    with driver.session() as s:
        rows = s.run(
            "MATCH (a:Authorization {dataset:$d}) "
            "WHERE apoc.any.property(a,'f_'+$field) IS NOT NULL "
            "UNWIND apoc.any.property(a,'f_'+$field) AS v "
            "WITH DISTINCT v WHERE v <> '*' "
            "RETURN v ORDER BY v LIMIT 500", d=dataset, field=field,
        )
        return [r["v"] for r in rows]


@app.post("/admin/org-profiles")
def admin_create_org_profile(req: OrgProfileCreateReq):
    name = req.name
    if not _SAFE_NAME.match(name):
        raise HTTPException(400, "ungueltiger Name (erlaubt: Buchstaben/Ziffern/._-)")
    cfg = profiles()
    if any(p["name"] == name for p in cfg["profiles"]):
        raise HTTPException(409, f"Org-Profil '{name}' existiert bereits")
    if not req.criteria:
        raise HTTPException(400, "mindestens ein Org-Kriterium erforderlich")
    custom_path = ensure_custom_org_profiles_file()
    custom = _load_json_list(custom_path)
    custom.append({"name": name, "description": req.description or "",
                   "org": {"mode": "filtered", "filters": _org_filters_from_criteria(req.criteria)}})
    custom_path.write_text(json.dumps(custom, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"name": name, "saved": True}


@app.put("/admin/org-profiles/{name}")
def admin_update_org_profile(name: str, req: OrgProfileEditReq):
    if name in PROTECTED_ORG_PROFILES:
        raise HTTPException(400, f"Org-Profil '{name}' ist geschuetzt und nicht editierbar")
    custom_path = ensure_custom_org_profiles_file()
    custom = _load_json_list(custom_path)
    entry = next((c for c in custom if c["name"] == name), None)
    if not entry:
        raise HTTPException(404, f"Org-Profil '{name}' nicht gefunden (oder ist ein Vendor-Profil)")
    if not req.criteria:
        raise HTTPException(400, "mindestens ein Org-Kriterium erforderlich")
    final_name = name
    new_name = (req.newName or "").strip()
    if new_name and new_name != name:
        if not _SAFE_NAME.match(new_name):
            raise HTTPException(400, "ungueltiger Name (erlaubt: Buchstaben/Ziffern/._-)")
        if new_name in PROTECTED_ORG_PROFILES:
            raise HTTPException(400, f"Name '{new_name}' ist reserviert (geschuetzte Basis-Variante)")
        if any(p["name"] == new_name for p in profiles()["profiles"]):
            raise HTTPException(409, f"Org-Profil '{new_name}' existiert bereits")
        entry["name"] = new_name
        final_name = new_name
    entry["description"] = req.description or ""
    entry["org"] = {"mode": "filtered", "filters": _org_filters_from_criteria(req.criteria)}
    custom_path.write_text(json.dumps(custom, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"name": final_name, "saved": True}


@app.delete("/admin/org-profiles/{name}")
def admin_delete_org_profile(name: str):
    if name in PROTECTED_ORG_PROFILES:
        raise HTTPException(400, f"Org-Profil '{name}' ist geschuetzt und nicht loeschbar")
    custom_path = ensure_custom_org_profiles_file()
    custom = _load_json_list(custom_path)
    remaining = [c for c in custom if c["name"] != name]
    if len(remaining) == len(custom):
        raise HTTPException(404, f"Org-Profil '{name}' nicht gefunden (oder ist ein Vendor-Profil)")
    custom_path.write_text(json.dumps(remaining, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"name": name, "deleted": True}


def _load_check_category(cat: str) -> list[dict]:
    path = CHECKS_DIR / f"{cat}.json"
    return json.loads(path.read_text(encoding="utf-8")) if path.is_file() else []


@app.get("/consistency-checks")
def consistency_checks(area: str = Query("user")):
    """Katalog der Konsistenzchecks (Qualitaet/Risiko des geladenen Berechtigungskonzepts,
    ruleset-unabhaengig) -- ein JSON je Kategorie unter checks/, Schema in checks/SCHEMA.md.
    Ausfuehrliche Begruendung je Check steht in KONSISTENZCHECKS.md (Quelle fuer Menschen).
    area='user' (Ribbon "User-spezifisch", Kategorien A-E) oder 'role' ("Rollen-spezifisch", R)."""
    cats = CHECK_AREAS.get(area, CHECK_AREAS["user"])
    catalog = []
    for cat in cats:
        catalog.extend(_load_check_category(cat))
    return catalog


def _find_check(check_id: str) -> dict:
    for cats in CHECK_AREAS.values():
        for cat in cats:
            for c in _load_check_category(cat):
                if c["id"] == check_id:
                    return c
    raise HTTPException(404, f"Check '{check_id}' nicht gefunden")


def run_cypher_path_capturing(session, path: Path, params: dict) -> list[list[dict]]:
    """Wie run_cypher_path, sammelt aber die Zeilen jedes Statements statt sie zu verwerfen --
    Konsistenzcheck-Dateien koennen mehrere ;-getrennte Statements haben (z. B. Zusammenfassung +
    Detailliste, s. sap_all.cypher); jedes Statement liefert ein eigenes Zeilen-Set."""
    return [[jsonable(dict(r)) for r in session.run(stmt, **params)]
            for stmt in split_statements(path.read_text(encoding="utf-8"))]


class ConsistencyRunReq(BaseModel):
    dataset: str
    # Kein asOf mehr hier: Stichtag kommt aus dem Dataset selbst (_dataset_asof()), s. RunReq.
    params: dict[str, int | str] = {}   # nur deklarierte Namen aus check["params"] wirksam (s. u.)


@app.post("/consistency-checks/{checkId}/run")
def run_consistency_check(checkId: str, req: ConsistencyRunReq):
    """Fuehrt die hinterlegte cypherFile eines Checks gegen ein Dataset aus (Stichtag = der
    Dataset-eigene asOf-Wert, s. _dataset_asof()) und gibt die Zeilen je Statement zurueck
    (results[0] = erstes Statement, ... ); nur fuer Checks mit implemented=true und gesetztem
    cypherFile (s. checks/SCHEMA.md). Optionale, vom Check deklarierte Schwellwerte (z. B.
    sleepDays bei B1) kommen aus check["params"] (Default) bzw. req.params (Override) -- nur
    deklarierte Namen werden als Cypher-Parameter durchgereicht."""
    check = _find_check(checkId)
    cypher_file = check.get("cypherFile")
    if not check.get("implemented") or not cypher_file:
        raise HTTPException(409, f"Check '{checkId}' ist noch nicht implementiert (kein Cypher vorhanden).")
    path = CYPHER_DIR / cypher_file.removeprefix("cypher/")
    if not path.is_file():
        raise HTTPException(500, f"Cypher-Datei fuer '{checkId}' fehlt: {cypher_file}")
    extra_params = {}
    for p in check.get("params", []):
        name = p["name"]
        extra_params[name] = req.params.get(name, p.get("default"))
    with driver.session() as s:
        as_of = _dataset_asof(s, req.dataset)
        results = run_cypher_path_capturing(s, path, {"dataset": req.dataset, "asOf": as_of, **extra_params})
    return {"checkId": checkId, "results": results}


# Anforderungstext je Befund fuer den A4-Root-Cause (s. critical_single_auths.cypher /
# critical_single_auths_root_cause.cypher) -- dieselben Kriterien, nur fuer die Anzeige
# aufbereitet (analog zur AuthReq-Anzeige im SoD-Root-Cause). Bei Aenderung der Kriterien dort
# auch hier nachziehen.
_A4_REQUIREMENT_TEXT = {
    "Debug-Replace (S_DEVELOP)": [
        {"field": "ACTVT", "andLogic": True, "values": ["02", "03"]},
        {"field": "OBJTYPE", "andLogic": True, "values": ["DEBUG"]},
    ],
    "Breiter Tabellenzugriff, aendern (S_TABU_DIS)": [
        {"field": "ACTVT", "andLogic": True, "values": ["02"]},
        {"field": "DICBERCLS", "andLogic": False, "values": ["*", "$"]},
    ],
    "Breiter Tabellenzugriff, aendern (S_TABU_NAM)": [
        {"field": "ACTVT", "andLogic": True, "values": ["02"]},
        {"field": "TABLE", "andLogic": True, "values": ["*"]},
    ],
    "Benutzergruppen-Verwaltung (S_USER_GRP)": [
        {"field": "(jede Ausprägung)", "andLogic": False, "values": ["*"]},
    ],
}


class ConsistencyRootCauseReq(BaseModel):
    dataset: str
    user: str


@app.post("/consistency-checks/{checkId}/root-cause")
def consistency_root_cause(checkId: str, req: ConsistencyRootCauseReq):
    """Root-Cause-Drilldown fuer einen einzelnen User innerhalb eines Konsistenzchecks -- zeigt,
    welche Rolle(n)/Profil(e) mit welchen konkreten Authorization-Feldwerten den Befund
    ausloesen. Nur fuer Checks mit gesetztem rootCauseFile (s. checks/SCHEMA.md). Antwortformat
    identisch zum SoD-Root-Cause (GET /root-cause, {blocks:[...]}), damit die UI denselben
    Dialog/dieselbe Render-Funktion wiederverwenden kann."""
    check = _find_check(checkId)
    rc_file = check.get("rootCauseFile")
    if not rc_file:
        raise HTTPException(409, f"Check '{checkId}' hat keinen Root-Cause-Drilldown.")
    path = CYPHER_DIR / rc_file.removeprefix("cypher/")
    if not path.is_file():
        raise HTTPException(500, f"Root-Cause-Cypher fuer '{checkId}' fehlt: {rc_file}")
    with driver.session() as s:
        as_of = _dataset_asof(s, req.dataset)
        stmt = split_statements(path.read_text(encoding="utf-8"))[0]
        rows = [jsonable(dict(r)) for r in s.run(stmt, dataset=req.dataset, asOf=as_of, user=req.user)]
    by_befund: dict[str, dict] = {}
    for r in rows:
        entry = by_befund.setdefault(r["befund"], {"object": r["objekt"], "satisfiedBy": []})
        felder_text = ", ".join(
            f"{k}: {','.join(v) if isinstance(v, list) else v}" for k, v in r["felder"].items())
        entry["satisfiedBy"].append({"actorType": r["akteurTyp"], "actorId": r["akteurId"],
                                      "authValues": [felder_text] if felder_text else []})
    blocks = [{"label": befund, "objects": [{
                  "object": data["object"],
                  "requirement": _A4_REQUIREMENT_TEXT.get(befund, []),
                  "satisfiedBy": data["satisfiedBy"],
               }]} for befund, data in by_befund.items()]
    return {"ruleId": checkId, "blocks": blocks}


class ConsistencyGraphReq(BaseModel):
    dataset: str


@app.post("/consistency-checks/{checkId}/graph")
def consistency_graph(checkId: str, req: ConsistencyGraphReq):
    """Graph-Ansicht eines Konsistenzchecks (Pilot fuer die Tabelle/Graph-Umschaltung, s.
    ROADMAP.md) -- nur fuer Checks mit gesetztem graphFile (s. checks/SCHEMA.md). Erwartet feste
    Spalten user/userType/userStatus/pathType/role/profile (role nullable bei direkter Zuweisung)
    und baut daraus generisch Cytoscape-Knoten/Kanten; dieselbe Spaltenform ist auf weitere
    User->(Rolle->)Profil-Checks uebertragbar, ohne den Konvertierungscode zu duplizieren."""
    check = _find_check(checkId)
    graph_file = check.get("graphFile")
    if not graph_file:
        raise HTTPException(409, f"Check '{checkId}' hat keine Graph-Ansicht.")
    path = CYPHER_DIR / graph_file.removeprefix("cypher/")
    if not path.is_file():
        raise HTTPException(500, f"Graph-Cypher fuer '{checkId}' fehlt: {graph_file}")
    with driver.session() as s:
        as_of = _dataset_asof(s, req.dataset)
        stmt = split_statements(path.read_text(encoding="utf-8"))[0]
        rows = [jsonable(dict(r)) for r in s.run(stmt, dataset=req.dataset, asOf=as_of)]
    nodes: dict[str, dict] = {}
    edges: dict[str, dict] = {}
    def add_node(node_id: str, label: str, kind: str, **extra):
        nodes.setdefault(node_id, {"data": {"id": node_id, "label": label, "kind": kind, **extra}})
    def add_edge(src: str, tgt: str, label: str):
        edges.setdefault(f"{src}->{tgt}", {"data": {"id": f"{src}->{tgt}", "source": src, "target": tgt, "label": label}})
    for r in rows:
        u_id, p_id = f"u:{r['user']}", f"p:{r['profile']}"
        add_node(u_id, r["user"], "User", userType=r["userType"], userStatus=r["userStatus"])
        add_node(p_id, r["profile"], "Profile")
        if r.get("role"):
            r_id = f"r:{r['role']}"
            add_node(r_id, r["role"], "Role")
            add_edge(u_id, r_id, "ASSIGNED_TO")
            add_edge(r_id, p_id, "HAS_PROFILE")
        else:
            add_edge(u_id, p_id, "HAS_PROFILE")
    return {"checkId": checkId, "elements": list(nodes.values()) + list(edges.values())}


def _collect_checks_data(session, dataset: str) -> tuple[str, list[list]]:
    """Fuehrt alle Konsistenzchecks mit Default-Params aus.
    Gibt (as_of_str, rows) zurueck; jede Row:
    [dataset, asOf, bereich, checkId, category, title, prio, status, treffer, params]
    wobei treffer=int|'' und status='implementiert'|'nicht implementiert'|'Fehler: ...'."""
    area_names = {"user": "User-spezifisch", "role": "Rollen-spezifisch", "import": "Import-spezifisch"}
    all_checks = []
    for area, cats in CHECK_AREAS.items():
        for cat in cats:
            for c in _load_check_category(cat):
                c["_area"] = area_names[area]
                all_checks.append(c)
    as_of = _dataset_asof(session, dataset)
    as_of_str = str(as_of) if as_of is not None else ""
    rows = []
    for c in all_checks:
        params_info = ""
        treffer: int | str = ""
        status = "nicht implementiert"
        if c.get("implemented") and c.get("cypherFile"):
            path = CYPHER_DIR / c["cypherFile"].removeprefix("cypher/")
            extra_params: dict = {}
            param_parts = []
            for p in c.get("params", []):
                v = p.get("default")
                extra_params[p["name"]] = v
                param_parts.append(f"{p['name']}={v}")
            params_info = "; ".join(param_parts)
            try:
                results = run_cypher_path_capturing(
                    session, path, {"dataset": dataset, "asOf": as_of, **extra_params})
                treffer = len(results[-1]) if results else 0
                status = "implementiert"
            except Exception as e:
                treffer = "Fehler"
                status = f"Fehler: {str(e)[:80]}"
        rows.append([
            dataset, as_of_str, c.get("_area", ""),
            c["id"], c.get("category", ""), c.get("title", ""),
            c.get("prio", ""), status, treffer, params_info,
        ])
    return as_of_str, rows


def _pdf_safe(text: str) -> str:
    """Ersetzt Zeichen ausserhalb Latin-1 durch ASCII-Aequivalente (fpdf2 built-in fonts)."""
    _MAP = {
        "„": '"', "“": '"', "”": '"',  # typografische Gaensefuesschen
        "‘": "'", "’": "'",                  # typografische Apostrophe
        "–": "-", "—": "-",                  # Gedankenstriche
        "…": "...",                                # Ellipsis
        "«": '"', "»": '"',                  # Guillemets
    }
    for src, dst in _MAP.items():
        text = text.replace(src, dst)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _build_consistency_pdf(
    dataset: str, as_of_str: str, rows: list[list],
    unternehmen: str, system: str, ersteller: str, anlass: str,
) -> bytes:
    """Baut den Konsistenz-Ueberblick-Report als PDF (fpdf2, Querformat A4)."""
    from fpdf import FPDF  # optional dep -- erst beim ersten PDF-Aufruf importiert

    NAVY = (28, 40, 82)
    PRIO_FG = {
        "Hoch":    (170,  20,  20),
        "Mittel":  (140,  70,   0),
        "Niedrig": ( 20, 110,  40),
        "Analytik":(  0,  70, 150),
    }
    PRIO_BG = {
        "Hoch":    (255, 222, 222),
        "Mittel":  (255, 244, 206),
        "Niedrig": (218, 248, 222),
        "Analytik":(207, 227, 255),
    }

    class _PDF(FPDF):
        def footer(self):
            self.set_y(-11)
            self.set_font("Helvetica", "I", 7)
            self.set_text_color(150, 150, 150)
            self.cell(
                0, 5,
                f"Vertraulich \xb7 Seite {self.page_no()}/{{nb}} \xb7 "
                "Erstellt mit IAM-Analysetool \xb7 Mandantendaten verbleiben in der lokalen Umgebung",
                align="C",
            )

    pdf = _PDF(orientation="L", format="A4")
    pdf.alias_nb_pages()
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()

    # ── Titelblock ───────────────────────────────────────────────────────────
    pdf.set_fill_color(*NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 12, "Konsistenzpr\xfcfung SAP-Berechtigungen", fill=True, ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 7,
             "Automatisierter Qualit\xe4ts- und Risikocheck des Berechtigungskonzepts",
             fill=True, ln=True)
    pdf.ln(5)

    # ── Stammdaten ───────────────────────────────────────────────────────────
    meta = [
        ("Unternehmen / Auftraggeber",   unternehmen or "-"),
        ("SAP-System / Mandant",          system      or "-"),
        ("Pr\xfcfungsanlass / Zeitraum", anlass      or "-"),
        ("Erstellt von",                 ersteller   or "-"),
        ("Dataset",                      dataset),
        ("Stichtag (Datenstand)",        as_of_str   or "-"),
        ("Erstellt am",                  datetime.date.today().isoformat()),
    ]
    for label, value in meta:
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(240, 242, 248)
        pdf.set_text_color(40, 50, 90)
        pdf.cell(65, 7, label, fill=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_fill_color(252, 252, 255)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(212, 7, _pdf_safe(str(value)[:90]), fill=True, ln=True)
    pdf.ln(6)

    # ── Abschnitts-Heading ───────────────────────────────────────────────────
    pdf.set_text_color(*NAVY)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Pr\xfcfergebnisse - \xdcbersicht", ln=True)
    pdf.ln(2)

    # ── Tabellen-Header ──────────────────────────────────────────────────────
    COL_W = [18, 153, 24, 22, 60]
    pdf.set_fill_color(*NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for w, h in zip(COL_W, ["ID", "Titel", "Priorit\xe4t", "Treffer", "Status"]):
        pdf.cell(w, 8, h, fill=True)
    pdf.ln(8)

    # ── Tabellenzeilen ───────────────────────────────────────────────────────
    current_area = None
    for i, row in enumerate(rows):
        bereich, check_id, title = row[2], row[3], row[5]
        prio, status_raw, treffer = row[6], row[7], row[8]

        if bereich != current_area:
            current_area = bereich
            pdf.set_fill_color(55, 70, 120)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(0, 6, _pdf_safe(f"  {bereich}"), fill=True, ln=True)

        is_impl = status_raw == "implementiert"
        is_err  = status_raw.startswith("Fehler")

        if is_impl:
            bg = PRIO_BG.get(prio, (250, 250, 250))
            if i % 2:
                bg = tuple(max(0, c - 8) for c in bg)
        elif is_err:
            bg = (255, 230, 230)
        else:
            bg = (236, 236, 240) if i % 2 else (244, 244, 248)

        pdf.set_fill_color(*bg)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 8)

        title_d   = _pdf_safe(title[:65] + "..." if len(title) > 68 else title)
        treffer_d = "-" if treffer == "" else ("!" if is_err else str(treffer))
        status_d  = ("implementiert" if is_impl
                     else "Fehler" if is_err else "nicht implementiert")

        pdf.cell(COL_W[0], 7, _pdf_safe(check_id), fill=True)
        pdf.cell(COL_W[1], 7, title_d,              fill=True)

        pdf.set_text_color(*(PRIO_FG.get(prio, (80, 80, 80)) if is_impl else (140, 140, 140)))
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(COL_W[2], 7, _pdf_safe(prio), fill=True)

        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(COL_W[3], 7, treffer_d, fill=True, align="R")
        pdf.cell(COL_W[4], 7, status_d,  fill=True, ln=True)

    # ── Zusammenfassung ──────────────────────────────────────────────────────
    impl   = [r for r in rows if r[7] == "implementiert"]
    total  = sum(r[8] for r in impl if isinstance(r[8], int))
    w_hits = sum(1 for r in impl if isinstance(r[8], int) and r[8] > 0)
    pdf.ln(4)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 5,
             f"{len(rows)} Checks gesamt \xb7 {len(impl)} implementiert \xb7 "
             f"{w_hits} mit Treffern \xb7 {total} Treffer insgesamt")

    return bytes(pdf.output())


@app.get("/consistency-checks/export")
def export_consistency_checks(dataset: str):
    """Konsistenz-Ueberblick-Report als CSV (Semikolon/UTF-8-BOM, Excel-tauglich)."""
    with driver.session() as s:
        as_of_str, rows = _collect_checks_data(s, dataset)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["Dataset", "Stichtag", "Bereich", "Check-ID", "Kategorie",
                "Titel", "Priorität", "Status", "Treffer", "Params"])
    for r in rows:
        w.writerow(r)
    data = "﻿" + buf.getvalue()
    fname = f"konsistenz_{dataset}_{as_of_str or 'unbekannt'}.csv"
    return Response(content=data, media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/consistency-checks/export/pdf")
def export_consistency_checks_pdf(
    dataset: str,
    unternehmen: str = Query(""),
    system: str = Query(""),
    ersteller: str = Query(""),
    anlass: str = Query(""),
):
    """Konsistenz-Ueberblick-Report als PDF (Querformat A4) mit Stammdaten-Kopfblock."""
    with driver.session() as s:
        as_of_str, rows = _collect_checks_data(s, dataset)
    try:
        pdf_bytes = _build_consistency_pdf(
            dataset, as_of_str, rows, unternehmen, system, ersteller, anlass)
    except ImportError:
        raise HTTPException(500, "fpdf2 nicht installiert — Image neu bauen: docker compose build backend")
    fname = f"konsistenz_{dataset}_{as_of_str or 'unbekannt'}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


def _build_import_evidence_pdf(
    dataset: str, as_of_str: str, summary: dict, rows: list[dict],
    unternehmen: str, system: str, ersteller: str, anlass: str,
) -> bytes:
    """Import-Evidenz-Report als PDF (fpdf2, Querformat A4) -- dediziertes Pendant zu
    _build_consistency_pdf() fuer die Pruefungsnachweisfuehrung: im allgemeinen Konsistenz-Report
    zaehlt I1 nur als ein Katalogeintrag mit Trefferzahl, hier steht die volle
    Tabelle-fuer-Tabelle-Rekonziliierung (Quellzeilen/Graph-Ergebnis/Status/Begruendung)."""
    from fpdf import FPDF  # optional dep -- erst beim ersten PDF-Aufruf importiert

    NAVY = (28, 40, 82)
    STATUS_FG = {
        "OK": (20, 110, 40), "Hinweis": (0, 70, 150),
        "Abweichung": (170, 20, 20), "Tabelle nicht im Extrakt (optional)": (110, 110, 110),
    }
    STATUS_BG = {
        "OK": (218, 248, 222), "Hinweis": (207, 227, 255),
        "Abweichung": (255, 222, 222), "Tabelle nicht im Extrakt (optional)": (236, 236, 240),
    }

    class _PDF(FPDF):
        def footer(self):
            self.set_y(-11)
            self.set_font("Helvetica", "I", 7)
            self.set_text_color(150, 150, 150)
            self.cell(
                0, 5,
                f"Vertraulich \xb7 Seite {self.page_no()}/{{nb}} \xb7 "
                "Erstellt mit IAM-Analysetool \xb7 Mandantendaten verbleiben in der lokalen Umgebung",
                align="C",
            )

    pdf = _PDF(orientation="L", format="A4")
    pdf.alias_nb_pages()
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()

    # ── Titelblock ───────────────────────────────────────────────────────────
    pdf.set_fill_color(*NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 12, "Import-Evidenz \xb7 Vollst\xe4ndigkeitsnachweis", fill=True, ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 7,
             "Quellzeilen je SAP-Tabelle gegen das Graph-Ergebnis abgeglichen (Pr\xfcfungsnachweisf\xfchrung)",
             fill=True, ln=True)
    pdf.ln(5)

    # ── Stammdaten ───────────────────────────────────────────────────────────
    meta = [
        ("Unternehmen / Auftraggeber",   unternehmen or "-"),
        ("SAP-System / Mandant",          system      or "-"),
        ("Pr\xfcfungsanlass / Zeitraum", anlass      or "-"),
        ("Erstellt von",                 ersteller   or "-"),
        ("Dataset",                      dataset),
        ("Stichtag (Datenstand)",        as_of_str   or "-"),
        ("Import-Zeitpunkt",             str(summary.get("importiertAm") or "-")),
        ("Erstellt am",                  datetime.date.today().isoformat()),
    ]
    for label, value in meta:
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(240, 242, 248)
        pdf.set_text_color(40, 50, 90)
        pdf.cell(65, 7, label, fill=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_fill_color(252, 252, 255)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(212, 7, _pdf_safe(str(value)[:90]), fill=True, ln=True)
    pdf.ln(6)

    # ── Zusammenfassung ──────────────────────────────────────────────────────
    pdf.set_text_color(*NAVY)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Zusammenfassung", ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 6,
             f"{summary.get('tabellenGeprueft', 0)} Tabellen gepr\xfcft \xb7 "
             f"{summary.get('ok', 0)} OK \xb7 {summary.get('hinweis', 0)} Hinweis \xb7 "
             f"{summary.get('nichtImExtrakt', 0)} nicht im Extrakt (optional) \xb7 "
             f"{summary.get('abweichung', 0)} Abweichung", ln=True)
    pdf.ln(4)

    # ── Tabellen-Header ──────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, "Rekonziliierung je Quelltabelle", ln=True)
    pdf.ln(1)
    COL_W = [24, 22, 26, 20, 18, 20, 20, 127]
    HEADERS = ["Tabelle", "Art", "Ziel", "Quellzeilen", "Gefiltert", "Graph", "Status", "Hinweis"]
    pdf.set_fill_color(*NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for w, h in zip(COL_W, HEADERS):
        pdf.cell(w, 8, h, fill=True)
    pdf.ln(8)

    # Kurzlabel fuers PDF -- "Tabelle nicht im Extrakt (optional)" sprengt jede sinnvolle
    # Spaltenbreite, im JSON/in der interaktiven Ansicht bleibt der volle Text.
    STATUS_SHORT = {"Tabelle nicht im Extrakt (optional)": "Optional"}

    def _cut(s: str, max_w: float) -> str:
        # Auf tatsaechliche Renderbreite (nicht Zeichenzahl) im gerade aktiven Font kuerzen --
        # fpdf2s cell() bricht/umbricht nicht selbst, zu langer Text wuerde sonst hart am
        # Zellenrand abgeschnitten (auch mitten im Wort, ohne Ellipse).
        s = _pdf_safe(str(s))
        if pdf.get_string_width(s) <= max_w:
            return s
        while s and pdf.get_string_width(s + "...") > max_w:
            s = s[:-1]
        return s + "..."

    # ── Tabellenzeilen ───────────────────────────────────────────────────────
    PAD = 2   # mm Innenabstand, damit der Text nicht bis an den Zellenrand reicht
    for i, r in enumerate(rows):
        status = r.get("status", "")
        bg = STATUS_BG.get(status, (244, 244, 248) if i % 2 else (236, 236, 240))
        fg = STATUS_FG.get(status, (0, 0, 0))
        dropped = r.get("verworfeneSpalten") or []
        hint = r.get("hinweis") or ""
        if dropped:
            hint = f"{hint} \xb7 verworfene Spalten: {', '.join(dropped)}"

        pdf.set_fill_color(*bg)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(COL_W[0], 7, _cut(r.get("tabelle") or "-", COL_W[0] - PAD), fill=True)
        pdf.cell(COL_W[1], 7, _cut(r.get("art") or "-", COL_W[1] - PAD), fill=True)
        pdf.cell(COL_W[2], 7, _cut(r.get("ziel") or "-", COL_W[2] - PAD), fill=True)
        pdf.cell(COL_W[3], 7, "-" if r.get("quellzeilen") is None else str(r["quellzeilen"]),
                 fill=True, align="R")
        pdf.cell(COL_W[4], 7, "-" if r.get("gefiltert") is None else str(r["gefiltert"]),
                 fill=True, align="R")
        pdf.cell(COL_W[5], 7, "-" if r.get("graphErgebnis") is None else str(r["graphErgebnis"]),
                 fill=True, align="R")
        pdf.set_text_color(*fg)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(COL_W[6], 7, _cut(STATUS_SHORT.get(status, status), COL_W[6] - PAD), fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(COL_W[7], 7, _cut(hint, COL_W[7] - PAD), fill=True, ln=True)

    return bytes(pdf.output())


@app.get("/datasets/{datasetId}/import-evidence/export")
def export_import_evidence_csv(datasetId: str):
    """Import-Evidenz-Rekonziliierung als CSV (Semikolon/UTF-8-BOM) -- Pendant zum PDF-Export
    fuer die Weiterverarbeitung (Excel/Revisionssoftware)."""
    with driver.session() as s:
        as_of = _dataset_asof(s, datasetId)
        results = run_cypher_path_capturing(
            s, CYPHER_DIR / "checks" / "import_evidence.cypher", {"dataset": datasetId, "asOf": as_of})
    rows = results[1] if len(results) > 1 else []
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["Tabelle", "Art", "Ziel", "Quellzeilen", "Gefiltert", "Verworfene Spalten",
                "Graph-Ergebnis", "Status", "Hinweis"])
    for r in rows:
        w.writerow([r.get("tabelle"), r.get("art"), r.get("ziel"), r.get("quellzeilen"),
                    r.get("gefiltert"), "|".join(r.get("verworfeneSpalten") or []),
                    r.get("graphErgebnis"), r.get("status"), r.get("hinweis")])
    data = "﻿" + buf.getvalue()
    fname = f"import_evidenz_{datasetId}_{as_of or 'unbekannt'}.csv"
    return Response(content=data, media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/datasets/{datasetId}/import-evidence")
def get_import_evidence(datasetId: str):
    """Import-Evidenz (Vollstaendigkeitsnachweis) des neuesten Imports dieses Datasets --
    Rekonziliierung Quellzeilen je SAP-Tabelle gegen das Graph-Ergebnis. Nutzt denselben
    Konsistenzcheck wie I1 (cypher/checks/import_evidence.cypher), hier als eigener,
    JSON-formfreundlicher Endpoint (Summary + Detail-Rekonziliierungsliste)."""
    with driver.session() as s:
        as_of = _dataset_asof(s, datasetId)
        results = run_cypher_path_capturing(
            s, CYPHER_DIR / "checks" / "import_evidence.cypher", {"dataset": datasetId, "asOf": as_of})
    return {
        "dataset": datasetId,
        "asOf": jsonable(as_of),
        "summary": jsonable(results[0][0]) if results[0] else {},
        "reconciliation": jsonable(results[1]),
    }


@app.get("/datasets/{datasetId}/import-evidence/export/pdf")
def export_import_evidence_pdf(
    datasetId: str,
    unternehmen: str = Query(""),
    system: str = Query(""),
    ersteller: str = Query(""),
    anlass: str = Query(""),
):
    """Import-Evidenz-Report als PDF (Querformat A4) mit voller Tabelle-fuer-Tabelle-
    Rekonziliierung -- fuer die Pruefungsnachweisfuehrung (im Unterschied zum allgemeinen
    Konsistenz-Report, der I1 nur als einen Katalogeintrag mit Trefferzahl zeigt)."""
    with driver.session() as s:
        as_of = _dataset_asof(s, datasetId)
        results = run_cypher_path_capturing(
            s, CYPHER_DIR / "checks" / "import_evidence.cypher", {"dataset": datasetId, "asOf": as_of})
    summary = results[0][0] if results[0] else {}
    rows = results[1]
    try:
        pdf_bytes = _build_import_evidence_pdf(
            datasetId, str(as_of) if as_of else "", summary, rows, unternehmen, system, ersteller, anlass)
    except ImportError:
        raise HTTPException(500, "fpdf2 nicht installiert — Image neu bauen: docker compose build backend")
    fname = f"import_evidenz_{datasetId}_{as_of or 'unbekannt'}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/admin/job-errors")
def admin_job_errors(limit: int = Query(200, ge=1, le=2000)):
    """Persistentes Fehlerprotokoll (Job-Fehler) ueber Container-Neustarts hinweg, neueste zuerst."""
    if not JOB_ERROR_LOG.is_file():
        return []
    entries = []
    for line in JOB_ERROR_LOG.read_text(encoding="utf-8").splitlines():
        try:
            entries.append(json.loads(line))
        except ValueError:
            continue
    return list(reversed(entries))[:limit]


@app.get("/admin/rulesets/{ruleset}/queries")
def admin_list_queries(ruleset: str):
    """Alle Queries eines Rulesets (Vendor + Overlay effektiv gemerged) fuer das Query
    Management; 'custom' markiert eigene Edits/abgeleitete Queries."""
    merged, custom_ids = _merged_queries(ruleset)
    return [{"id": qid, "description": q.get("description", ""),
             "shortDescription": q.get("shortDescription", ""), "criticality": q.get("criticality"),
             "criticalityRank": q.get("criticalityRank", 0),
             "module": q.get("module"), "queryType": q.get("queryType"),
             "disregardTcode": bool(q.get("disregardTcode", False)), "custom": qid in custom_ids}
            for qid, q in sorted(merged.items())]


@app.get("/admin/rulesets/{ruleset}/queries/{queryId}")
def admin_get_query(ruleset: str, queryId: str):
    """Eine Query vollstaendig (inkl. authorizations/transactions, read-only) fuer den
    'Aufbau'-Tab im Query Management."""
    merged, custom_ids = _merged_queries(ruleset)
    q = merged.get(queryId)
    if not q:
        raise HTTPException(404, f"Query '{queryId}' nicht gefunden")
    return {**q, "custom": queryId in custom_ids}


@app.get("/admin/rulesets/{ruleset}/overlay/download")
def admin_download_overlay(ruleset: str):
    """Overlay-Datei (queries.custom.json) eines Rulesets als Download — Sicherung der eigenen
    Anpassungen/abgeleiteten Queries, getrennt von den Quelldaten-/Lauf-Backups."""
    custom_path = ensure_custom_queries_file(ruleset)
    return FileResponse(custom_path, filename=f"{ruleset}__queries.custom.json", media_type="application/json")


class QueryEditReq(BaseModel):
    description: str | None = None
    shortDescription: str | None = None
    criticality: str | None = None
    module: str | None = None
    queryType: str | None = None
    disregardTcode: bool | None = None
    risk: str | None = None
    controls: str | None = None


@app.put("/admin/rulesets/{ruleset}/queries/{queryId}")
def admin_update_query(ruleset: str, queryId: str, req: QueryEditReq):
    """Bearbeitet Metadaten einer Query als Overlay-Eintrag (queries.custom.json) — Vendor-Datei
    bleibt unberuehrt. Ladet das Ruleset danach sofort neu (Edit wirkt ohne extra Schritt)."""
    merged, _ = _merged_queries(ruleset)
    if queryId not in merged:
        raise HTTPException(404, f"Query '{queryId}' nicht gefunden")
    custom_path = ensure_custom_queries_file(ruleset)
    custom = _load_json_list(custom_path)
    entry = next((c for c in custom if c["query"] == queryId), None)
    fields = {k: v for k, v in req.model_dump().items() if v is not None}
    if entry:
        entry.update(fields)
    else:
        custom.append({"query": queryId, **fields})
    custom_path.write_text(json.dumps(custom, ensure_ascii=False, indent=2), encoding="utf-8")
    reload_ruleset(ruleset)
    return {"query": queryId, "saved": fields}


class QueryDeriveReq(BaseModel):
    newId: str
    fromId: str
    description: str | None = None
    shortDescription: str | None = None
    criticality: str | None = None
    module: str | None = None
    queryType: str | None = None
    disregardTcode: bool | None = None
    risk: str | None = None
    controls: str | None = None


@app.post("/admin/rulesets/{ruleset}/queries/derive")
def admin_derive_query(ruleset: str, req: QueryDeriveReq):
    """Legt eine neue Query als Kopie einer bestehenden an (authorizations/transactions 1:1
    uebernommen; nur Metadaten hier optional ueberschrieben) — landet im Overlay, die
    Quell-Query bleibt unberuehrt. Ladet das Ruleset danach sofort neu."""
    if not _SAFE_NAME.match(req.newId):
        raise HTTPException(400, "ungueltige Query-ID (erlaubt: Buchstaben/Ziffern/._-)")
    merged, _ = _merged_queries(ruleset)
    if req.newId in merged:
        raise HTTPException(409, f"Query-ID '{req.newId}' existiert bereits")
    src = merged.get(req.fromId)
    if not src:
        raise HTTPException(404, f"Quell-Query '{req.fromId}' nicht gefunden")
    new_q = dict(src)
    new_q["query"] = req.newId
    new_q["derivedFrom"] = req.fromId
    for field in ("description", "shortDescription", "criticality", "module", "queryType",
                  "disregardTcode", "risk", "controls"):
        v = getattr(req, field)
        if v is not None:
            new_q[field] = v
    custom_path = ensure_custom_queries_file(ruleset)
    custom = _load_json_list(custom_path)
    custom.append(new_q)
    custom_path.write_text(json.dumps(custom, ensure_ascii=False, indent=2), encoding="utf-8")
    reload_ruleset(ruleset)
    return {"query": req.newId, "derivedFrom": req.fromId}


# --- SoD-Regel-Editor (Query Management, Modus "SoD") ----------------------------------
# Analog zum Einzelfilter-Editor oben, aber ohne Struktur-Edits/Ableiten: SoD-Regeln haben keine
# eigene Authorizations/TCodes, nur Metadaten (Kurz-/Langbezeichnung, Kritikalitaet, Risiko,
# Controls) plus die read-only Klausel-/Variablen-Struktur (clauses/variables/expression) fuer den
# Aufbau-Tab. Overlay sod_rules.custom.json analog zu queries.custom.json.
def _sodrule_paths(ruleset: str) -> tuple[Path, Path]:
    rdir = ruleset_dir(ruleset)
    if not rdir:
        raise HTTPException(404, f"Ruleset '{ruleset}' nicht gefunden")
    base = RULES_DIR / rdir
    return base / "sod_rules.json", base / "sod_rules.custom.json"


def ensure_custom_sodrules_file(ruleset: str) -> Path:
    _, custom_path = _sodrule_paths(ruleset)
    if not custom_path.is_file():
        custom_path.write_text("[]", encoding="utf-8")
    return custom_path


def _merged_sodrules(ruleset: str) -> tuple[dict[str, dict], set[str]]:
    """Vendor-SoD-Regeln + Overlay, Overlay-Felder gewinnen je id — analog zu _merged_queries.
    'eigen' heisst: neue Regel (kein Vendor-Gegenstueck) oder clauses/variables wurden
    ueberschrieben; reine Metadaten-Edits zaehlen nicht."""
    vendor_path, custom_path = _sodrule_paths(ruleset)
    merged = {r["sodRule"]: dict(r) for r in _load_json_list(vendor_path)}
    custom_ids = set()
    for c in _load_json_list(custom_path):
        rid = c["sodRule"]
        if rid not in merged or "clauses" in c or "variables" in c:
            custom_ids.add(rid)
        if rid in merged:
            merged[rid] = {**merged[rid], **{k: v for k, v in c.items() if v is not None}}
        else:
            merged[rid] = c
    return merged, custom_ids


@app.get("/admin/rulesets/{ruleset}/sodrules")
def admin_list_sodrules(ruleset: str):
    """Alle SoD-Regeln eines Rulesets (Vendor + Overlay gemerged) fuer das Query Management
    (Modus 'SoD'); 'custom' markiert eigene Edits. 'clauses' (CNF: UND ueber Klauseln, ODER
    innerhalb einer Klausel -- nur bei Rulesets mit CNF-Struktur, aktuell nur kpmg_r3, sonst [])
    speist die clientseitige bidirektionale Einzelfilter<->SoD-Verknuepfung der Katalog-Auswahl
    (Assistent Schritt Scoping / Admin-Seite "Scope")."""
    merged, custom_ids = _merged_sodrules(ruleset)
    return [{"id": rid, "description": r.get("description", ""),
             "shortDescription": r.get("shortDescription", ""), "criticality": r.get("criticality"),
             "criticalityRank": r.get("criticalityRank", 0), "clauses": r.get("clauses", []),
             "reasonCode": r.get("reasonCode"), "custom": rid in custom_ids}
            for rid, r in sorted(merged.items())]


@app.get("/admin/rulesets/{ruleset}/sodrules/{ruleId}")
def admin_get_sodrule(ruleset: str, ruleId: str):
    """Eine SoD-Regel vollstaendig (inkl. clauses/variables/expression, read-only) fuer den
    'Aufbau'-Tab im Query Management."""
    merged, custom_ids = _merged_sodrules(ruleset)
    r = merged.get(ruleId)
    if not r:
        raise HTTPException(404, f"SoD-Regel '{ruleId}' nicht gefunden")
    return {**r, "custom": ruleId in custom_ids}


@app.get("/admin/rulesets/{ruleset}/sodrules/overlay/download")
def admin_download_sodrule_overlay(ruleset: str):
    """Overlay-Datei (sod_rules.custom.json) eines Rulesets als Download."""
    custom_path = ensure_custom_sodrules_file(ruleset)
    return FileResponse(custom_path, filename=f"{ruleset}__sod_rules.custom.json", media_type="application/json")


class SodRuleEditReq(BaseModel):
    description: str | None = None
    shortDescription: str | None = None
    criticality: str | None = None
    risk: str | None = None
    controls: str | None = None


@app.put("/admin/rulesets/{ruleset}/sodrules/{ruleId}")
def admin_update_sodrule(ruleset: str, ruleId: str, req: SodRuleEditReq):
    """Bearbeitet Metadaten einer SoD-Regel als Overlay-Eintrag (sod_rules.custom.json) — Vendor-
    Datei bleibt unberuehrt. Ladet das Ruleset danach sofort neu (Edit wirkt ohne extra Schritt)."""
    merged, _ = _merged_sodrules(ruleset)
    if ruleId not in merged:
        raise HTTPException(404, f"SoD-Regel '{ruleId}' nicht gefunden")
    custom_path = ensure_custom_sodrules_file(ruleset)
    custom = _load_json_list(custom_path)
    entry = next((c for c in custom if c["sodRule"] == ruleId), None)
    fields = {k: v for k, v in req.model_dump().items() if v is not None}
    if entry:
        entry.update(fields)
    else:
        custom.append({"sodRule": ruleId, **fields})
    custom_path.write_text(json.dumps(custom, ensure_ascii=False, indent=2), encoding="utf-8")
    reload_ruleset(ruleset)
    return {"sodRule": ruleId, "saved": fields}


# --- Scope-Profile (persistente Katalog-Auswahl, Admin-Seite "Scope") -------------------
# Analog zu den Query-/SoD-Regel-Overlays: reine Custom-Datei je Ruleset-Ordner (kein Vendor-
# Gegenstueck noetig), git-getrackt wie queries.custom.json/sod_rules.custom.json -- anders als
# die Org-Varianten (config/analysis_profiles.custom.json, bewusst .gitignore'd) enthalten
# Query-/SoD-Regel-IDs keine Mandantendaten. Ersetzt funktional das nie verdrahtete Phase-3-
# Scaffold "scopeProfiles" in config/analysis_profiles.json (bleibt unangetastet stehen).
def _scope_profiles_path(ruleset: str) -> Path:
    rdir = ruleset_dir(ruleset)
    if not rdir:
        raise HTTPException(404, f"Ruleset '{ruleset}' nicht gefunden")
    return RULES_DIR / rdir / "scope_profiles.custom.json"


def ensure_custom_scope_profiles_file(ruleset: str) -> Path:
    path = _scope_profiles_path(ruleset)
    if not path.is_file():
        path.write_text("[]", encoding="utf-8")
    return path


class ScopeProfileEditReq(BaseModel):
    description: str | None = None
    queryIds: list[str] = []
    sodRuleIds: list[str] = []
    # Voreinstellung fuer "Neuer Lauf": ueberschreibt dort userTypeProfile/sleepDays, wenn dieser
    # Scope gewaehlt wird (s. currentRunScopingSource() im Frontend). Defaults hier nur Fallback
    # fuer direkte API-Aufrufe ohne die Felder -- das Frontend befuellt sie beim Anlegen immer
    # aus den echten aktuellen /profiles-Werten.
    userTypeProfile: str = "all"
    sleepDays: int = 180


class ScopeProfileCreateReq(ScopeProfileEditReq):
    name: str


def _scope_profile_nonempty(req: ScopeProfileEditReq) -> None:
    if not req.queryIds and not req.sodRuleIds:
        raise HTTPException(400, "mindestens ein Einzelfilter oder eine SoD-Regel erforderlich")


@app.get("/admin/rulesets/{ruleset}/scopes")
def admin_list_scopes(ruleset: str):
    """Gespeicherte Scope-Profile eines Rulesets (Katalog-Auswahl, wiederverwendbar ueber
    Datasets hinweg) -- speist die Admin-Seite 'Scope' und die Auswahl im 'Neuer Lauf'-Dialog."""
    return _load_json_list(_scope_profiles_path(ruleset))


@app.post("/admin/rulesets/{ruleset}/scopes")
def admin_create_scope(ruleset: str, req: ScopeProfileCreateReq):
    if not _SAFE_NAME.match(req.name):
        raise HTTPException(400, "ungueltiger Name (erlaubt: Buchstaben/Ziffern/._-)")
    _scope_profile_nonempty(req)
    path = ensure_custom_scope_profiles_file(ruleset)
    scopes = _load_json_list(path)
    if any(s["name"] == req.name for s in scopes):
        raise HTTPException(409, f"Scope-Profil '{req.name}' existiert bereits")
    scopes.append({"name": req.name, "description": req.description or "",
                   "queryIds": req.queryIds, "sodRuleIds": req.sodRuleIds,
                   "userTypeProfile": req.userTypeProfile, "sleepDays": req.sleepDays})
    path.write_text(json.dumps(scopes, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"name": req.name, "saved": True}


@app.put("/admin/rulesets/{ruleset}/scopes/{name}")
def admin_update_scope(ruleset: str, name: str, req: ScopeProfileEditReq):
    _scope_profile_nonempty(req)
    path = ensure_custom_scope_profiles_file(ruleset)
    scopes = _load_json_list(path)
    entry = next((s for s in scopes if s["name"] == name), None)
    if not entry:
        raise HTTPException(404, f"Scope-Profil '{name}' nicht gefunden")
    entry["description"] = req.description or ""
    entry["queryIds"] = req.queryIds
    entry["sodRuleIds"] = req.sodRuleIds
    entry["userTypeProfile"] = req.userTypeProfile
    entry["sleepDays"] = req.sleepDays
    path.write_text(json.dumps(scopes, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"name": name, "saved": True}


@app.delete("/admin/rulesets/{ruleset}/scopes/{name}")
def admin_delete_scope(ruleset: str, name: str):
    path = ensure_custom_scope_profiles_file(ruleset)
    scopes = _load_json_list(path)
    remaining = [s for s in scopes if s["name"] != name]
    if len(remaining) == len(scopes):
        raise HTTPException(404, f"Scope-Profil '{name}' nicht gefunden")
    path.write_text(json.dumps(remaining, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"name": name, "deleted": True}


@app.get("/runs")
def list_runs():
    with driver.session() as s:
        out = []
        for r in s.run(
                "MATCH (run:Run) "
                "CALL { WITH run MATCH (f:SoDConflict {runId:run.runId, dataset:run.dataset}) "
                "  RETURN count(f) AS findings, count(DISTINCT f.ruleId) AS rules, "
                "         sum(CASE WHEN f.userSleeping THEN 1 ELSE 0 END) AS sleeping } "
                "RETURN run, findings, rules, sleeping ORDER BY run.runId"):
            d = jsonable(dict(r["run"]))
            d.update(findings=r["findings"], rules=r["rules"], sleeping=r["sleeping"])
            out.append(d)
        return out


_FINDINGS_WHERE = (
    # effSleeping/effLastLogonKnown: ohne $sleepDaysOverride die beim Lauf materialisierten Werte
    # (f.userSleeping/f.lastLogonKnown, fest an run.sleepDays gebunden) -- mit Override live gegen
    # u.lastLogon/run.asOf gerechnet (identische Formel wie _USER_ENRICH_RETURN), damit die
    # Sleeping-Schnellwahl (90/180/360 Tage) im Ergebnisfilter unabhaengig vom beim Lauf gewaehlten
    # Fenster funktioniert, ohne den Lauf neu zu materialisieren.
    "WITH u, f, rule, run, "
    "     CASE WHEN $sleepDaysOverride IS NULL THEN coalesce(f.userSleeping,false) "
    "          ELSE (u.lastLogon IS NOT NULL AND u.lastLogon < (run.asOf - duration({days: $sleepDaysOverride}))) END AS effSleeping, "
    "     CASE WHEN $sleepDaysOverride IS NULL THEN coalesce(f.lastLogonKnown,true) "
    "          ELSE (u.lastLogon IS NOT NULL) END AS effLastLogonKnown "
    "WHERE coalesce(f.criticalityRank,0) >= $minRank "
    "  AND ($user IS NULL OR u.id = $user) "
    "  AND ($rule IS NULL OR f.ruleId = $rule) "
    "  AND (size($userTypes) = 0 OR any(t IN $userTypes WHERE t IN labels(u))) "
    "  AND ($sleeping IS NULL "
    "       OR ($sleeping = 'true' AND effSleeping = true) "
    "       OR ($sleeping = 'false' AND effSleeping = false AND effLastLogonKnown = true) "
    "       OR ($sleeping = 'unknown' AND effLastLogonKnown = false)) "
    "  AND ($ruleCriticality IS NULL OR f.criticality = $ruleCriticality) "
    # locked/lockReason: Sperrstatus/-grund liegen direkt am User (u.lockReasons, aus dem UFLAG-
    # Bitfeld, s. load/01_users.cypher) -- kein materialisiertes Aequivalent am Finding noetig.
    # Funktioniert nur fuer Laeufe, die gesperrte User nicht schon beim Materialisieren ausgeschlossen
    # haben (excludeLocked=false, s. userTypeProfile) -- sonst existieren dafuer schlicht keine
    # Findings.
    "  AND ($locked IS NULL "
    "       OR ($locked = 'true' AND 'Locked' IN labels(u) "
    "           AND ($lockReason IS NULL OR $lockReason = 'alle' OR $lockReason IN coalesce(u.lockReasons,[]))) "
    "       OR ($locked = 'false' AND NOT 'Locked' IN labels(u))) "
)


@app.get("/findings")
def findings(runId: str, minRank: int = 0, limit: int = 200,
             user: str | None = None, rule: str | None = None,
             userType: list[str] = Query(default=[]), sleeping: str | None = None,
             ruleCriticality: str | None = None, sleepDaysOverride: int | None = None,
             locked: str | None = None, lockReason: str | None = None):
    """Findings eines Laufs; optional auf User/Regel/Nutzertyp(en)/Sleeping/Kritikalitaet/
    Sperrstatus eingeschraenkt (Drill-down: Klick auf User-/Regel-Zelle, oder Ergebnisfilter in
    der Sidebar). sleeping: 'true' (bestaetigt sleeping) / 'false' (bestaetigt aktiv) / 'unknown'
    (kein TRDAT). sleepDaysOverride: weicht vom beim Lauf gesetzten sleepDays-Fenster ab (Sleeping-
    Schnellwahl 90/180/360) -- schaltet die Sleeping-Berechnung von materialisiert auf live gegen
    u.lastLogon um. locked: 'true'/'false'; lockReason: 'alle'/'failed_logons'/'admin_local'/
    'admin_global' (nur wirksam bei locked='true')."""
    with driver.session() as s:
        return [jsonable(dict(r)) for r in s.run(
            "MATCH (u:User)-[:VIOLATES]->(f:SoDConflict {runId:$runId})-[:BASED_ON]->(rule:SoDRule) "
            "MATCH (run:Run {runId:$runId}) "
            + _FINDINGS_WHERE +
            "RETURN u.id AS user, f.ruleId AS rule, f.criticality AS criticality, "
            "       effSleeping AS sleeping, effLastLogonKnown AS lastLogonKnown, "
            "       f.conflictType AS conflictType, "
            "       [(f)-[:VIA_ROLE]->(r) | r.id] AS roles, "
            "       [(f)-[:VIA_PROFILE]->(p) | p.id] AS profiles "
            "ORDER BY coalesce(f.criticalityRank,0) DESC, user LIMIT $limit",
            runId=runId, minRank=minRank, limit=limit, user=user, rule=rule,
            userTypes=userType, sleeping=sleeping, ruleCriticality=ruleCriticality,
            sleepDaysOverride=sleepDaysOverride, locked=locked, lockReason=lockReason)]


@app.get("/findings/summary")
def findings_summary(runId: str, minRank: int = 0, user: str | None = None, rule: str | None = None,
                      userType: list[str] = Query(default=[]), sleeping: str | None = None,
                      ruleCriticality: str | None = None, sleepDaysOverride: int | None = None,
                      locked: str | None = None, lockReason: str | None = None):
    """KPI-Aggregate (Findings/betroffene Regeln/sleeping/unbekannt) fuer den aktuellen
    Filterkontext — unabhaengig vom Limit der Findings-Liste, damit die KPI-Kacheln zum
    gewaehlten Filter passen. Parameter s. GET /findings."""
    with driver.session() as s:
        rec = s.run(
            "MATCH (u:User)-[:VIOLATES]->(f:SoDConflict {runId:$runId})-[:BASED_ON]->(rule:SoDRule) "
            "MATCH (run:Run {runId:$runId}) "
            + _FINDINGS_WHERE +
            "RETURN count(f) AS findings, count(DISTINCT f.ruleId) AS rules, "
            "       sum(CASE WHEN effSleeping THEN 1 ELSE 0 END) AS sleeping, "
            "       sum(CASE WHEN NOT effLastLogonKnown THEN 1 ELSE 0 END) AS unknownLogon",
            runId=runId, minRank=minRank, user=user, rule=rule,
            userTypes=userType, sleeping=sleeping, ruleCriticality=ruleCriticality,
            sleepDaysOverride=sleepDaysOverride, locked=locked, lockReason=lockReason).single()
        return jsonable(dict(rec))


def _query_scope_where(query_ids: list[str], sod_rules: list[str], query_scope: str) -> str:
    """WHERE-Fragment fuer 'welche Queries wurden/werden in diesem Lauf betrachtet' -- muss exakt
    widerspiegeln, was materialize_matches_candidates.cypher fuer den Query-Umfang eines Laufs
    auswaehlt, gleiche Prioritaet: explizite $queryIds (Katalog-Auswahl) > $sodRules-Scoping (nur
    Klausel-Queries dieser Regeln) > altes $queryScope ('all' -> jede Query des Rulesets,
    'sodOnly' -> nur als SoD-Klausel verbaute Queries; Default fuer Laeufe ohne gespeichertes
    queryScope). Aeltere Laeufe ohne queryIds/sodRules (coalesce -> []) fallen automatisch auf
    das bisherige queryScope-Verhalten zurueck."""
    if query_ids:
        return "WHERE q.id IN $queryIds "
    if sod_rules:
        return ("WHERE EXISTS { MATCH (q)<-[:NEEDS]-(:Clause {ruleset:$ruleset})<-[:HAS_CLAUSE]-(rule:SoDRule {ruleset:$ruleset}) "
                 "WHERE rule.id IN $sodRules } ")
    return "" if query_scope == "all" else "WHERE EXISTS { (q)<-[:NEEDS]-(:Clause {ruleset:$ruleset}) } "


@app.get("/queries")
def queries(runId: str):
    """Queries (Einzelfilter) eines Laufs, beschraenkt auf dessen tatsaechlichen Query-Umfang
    (s. r.queryScope/r.queryIds/r.sodRules + _query_scope_where) — speist die Query-Auswahl fuer
    den Matches-Drill-down ('wer matcht Query X') und die Sidebar-Filter der Ergebnis-Ansicht."""
    with driver.session() as s:
        run = s.run(
            "MATCH (r:Run {runId:$id}) RETURN r.ruleset AS ruleset, coalesce(r.queryScope,'sodOnly') AS queryScope, "
            "coalesce(r.queryIds,[]) AS queryIds, coalesce(r.sodRules,[]) AS sodRules",
            id=runId).single()
        if not run:
            raise HTTPException(404, f"Lauf '{runId}' nicht gefunden")
        return [dict(r) for r in s.run(
            "MATCH (q:Query {ruleset:$ruleset}) "
            + _query_scope_where(run["queryIds"], run["sodRules"], run["queryScope"]) +
            "RETURN q.id AS id, q.description AS description, q.shortDescription AS shortDescription, "
            "q.criticality AS criticality, q.module AS module ORDER BY id",
            ruleset=run["ruleset"], queryIds=run["queryIds"], sodRules=run["sodRules"])]


@app.get("/sodrules")
def sod_rules(runId: str):
    """SoD-Regeln (mit Bezeichnung) eines Laufs — speist die SoD-Auswahl in der Sidebar
    (Bezeichnung statt nur der Regel-ID). Wurden beim Lauf explizit SoD-Regeln ausgewaehlt
    (Katalog-Auswahl-Scope, r.sodRules), zeigt die Liste nur diese; sonst (queryScope-basierter
    oder alter Lauf ohne r.sodRules) weiterhin alle SoD-Regeln des Rulesets."""
    with driver.session() as s:
        run = s.run(
            "MATCH (r:Run {runId:$id}) RETURN r.ruleset AS ruleset, coalesce(r.sodRules,[]) AS sodRules",
            id=runId).single()
        if not run:
            raise HTTPException(404, f"Lauf '{runId}' nicht gefunden")
        where = "WHERE rule.id IN $sodRules " if run["sodRules"] else ""
        return [dict(r) for r in s.run(
            "MATCH (rule:SoDRule {ruleset:$ruleset}) " + where +
            "RETURN rule.id AS id, rule.description AS description, "
            "rule.shortDescription AS shortDescription, rule.criticality AS criticality "
            "ORDER BY id", ruleset=run["ruleset"], sodRules=run["sodRules"])]


@app.get("/queries/summary")
def queries_summary(runId: str):
    """Einzelberechtigungs-Uebersicht (Ergebnisse-Menue): pro Query, wie viele User sie in diesem
    Lauf matchen. Nur Queries mit mindestens einem Treffer (0-Treffer waere Katalog-Browsing, kein
    Ergebnis -- die MATCH-Kardinalitaet filtert das implizit; welche Queries ueberhaupt eine
    MATCHES-Kante bekommen konnten, entscheidet bereits der Query-Umfang des Laufs, s. r.queryScope
    in materialize_matches_candidates.cypher -- hier also KEIN zusaetzlicher Klausel-Filter noetig).
    Klick auf eine Zeile filtert im Frontend die normale Einzelfilter-Ansicht auf diese Query.
    totalUsers = distinkte User ueber ALLE Zeilen zusammen (fuer die Kopf-Kachel; naives Aufsummieren
    von userCount je Zeile waere falsch, da ein User i. d. R. mehrere Queries matcht)."""
    with driver.session() as s:
        run = s.run("MATCH (r:Run {runId:$id}) RETURN r.ruleset AS ruleset", id=runId).single()
        if not run:
            raise HTTPException(404, f"Lauf '{runId}' nicht gefunden")
        ruleset = run["ruleset"]
        rows = [dict(r) for r in s.run(
            "MATCH (u:User)-[:MATCHES {ruleset:$ruleset, runId:$runId}]->(q:Query {ruleset:$ruleset}) "
            "WITH q, count(DISTINCT u) AS userCount "
            "RETURN q.id AS id, q.description AS description, q.shortDescription AS shortDescription, "
            "  q.criticality AS criticality, coalesce(q.criticalityRank,0) AS criticalityRank, "
            "  q.module AS module, userCount "
            "ORDER BY coalesce(q.criticalityRank,0) DESC, userCount DESC",
            ruleset=ruleset, runId=runId)]
        total_users = s.run(
            "MATCH (u:User)-[:MATCHES {ruleset:$ruleset, runId:$runId}]->(:Query {ruleset:$ruleset}) "
            "RETURN count(DISTINCT u) AS c", ruleset=ruleset, runId=runId).single()["c"]
        return {"totalUsers": total_users, "rows": rows}


@app.get("/sodrules/summary")
def sod_rules_summary(runId: str):
    """SoD-Regel-Uebersicht (Ergebnisse-Menue): pro Regel, wie viele User sie in diesem Lauf
    verletzen. Nur Regeln mit mindestens einem Fund. Klick auf eine Zeile filtert im Frontend
    die normale Findings-Liste auf diese Regel. totalUsers s. queries_summary."""
    with driver.session() as s:
        run = s.run("MATCH (r:Run {runId:$id}) RETURN r.ruleset AS ruleset", id=runId).single()
        if not run:
            raise HTTPException(404, f"Lauf '{runId}' nicht gefunden")
        ruleset = run["ruleset"]
        rows = [dict(r) for r in s.run(
            "MATCH (rule:SoDRule {ruleset:$ruleset}) "
            "WHERE EXISTS { (rule)-[:HAS_CLAUSE]->() } "
            "MATCH (u:User)-[:VIOLATES]->(:SoDConflict {ruleset:$ruleset, runId:$runId, ruleId:rule.id}) "
            "WITH rule, count(DISTINCT u) AS userCount "
            "RETURN rule.id AS id, rule.description AS description, "
            "  rule.shortDescription AS shortDescription, rule.criticality AS criticality, "
            "  coalesce(rule.criticalityRank,0) AS criticalityRank, userCount "
            "ORDER BY coalesce(rule.criticalityRank,0) DESC, userCount DESC",
            ruleset=ruleset, runId=runId)]
        total_users = s.run(
            "MATCH (u:User)-[:VIOLATES]->(:SoDConflict {ruleset:$ruleset, runId:$runId}) "
            "RETURN count(DISTINCT u) AS c", ruleset=ruleset, runId=runId).single()["c"]
        return {"totalUsers": total_users, "rows": rows}


_USER_TYP_STATUS_CYPHER = (
    "CASE WHEN 'Dialog' IN labels(u) THEN 'Dialog' WHEN 'System' IN labels(u) THEN 'System' "
    "     WHEN 'Service' IN labels(u) THEN 'Service' WHEN 'Communication' IN labels(u) THEN 'Communication' "
    "     WHEN 'Reference' IN labels(u) THEN 'Reference' ELSE '?' END AS typ, "
    "CASE WHEN 'Locked' IN labels(u) THEN 'gesperrt' ELSE 'aktiv' END AS status"
)


def _summary_export_rows(s, kind: str, run_id: str, detailed: bool):
    """Liefert die Zeilen fuer den Ergebnisse-Uebersicht-Export (Einzelfilter/SoD) — dieselbe
    Aggregation wie GET /queries/summary bzw. /sodrules/summary, optional (detailed=True, fuer
    'ausfuehrliches Excel') je Zeile zusaetzlich die Liste der betroffenen Nutzer (Grundlage fuer
    die Excel-Gruppierung/Auffaltung)."""
    run = s.run("MATCH (r:Run {runId:$id}) RETURN r.ruleset AS ruleset", id=run_id).single()
    if not run:
        raise HTTPException(404, f"Lauf '{run_id}' nicht gefunden")
    ruleset = run["ruleset"]
    group_by = "q" if kind == "query" else "rule"
    match_clause = (
        "MATCH (u:User)-[:MATCHES {ruleset:$ruleset, runId:$runId}]->(q:Query {ruleset:$ruleset}) "
        if kind == "query" else
        "MATCH (rule:SoDRule {ruleset:$ruleset}) WHERE EXISTS { (rule)-[:HAS_CLAUSE]->() } "
        "MATCH (u:User)-[:VIOLATES]->(:SoDConflict {ruleset:$ruleset, runId:$runId, ruleId:rule.id}) "
    )
    return_fields = (
        "q.id AS id, q.description AS description, q.shortDescription AS shortDescription, "
        "q.criticality AS criticality, coalesce(q.criticalityRank,0) AS criticalityRank, q.module AS module"
        if kind == "query" else
        "rule.id AS id, rule.description AS description, rule.shortDescription AS shortDescription, "
        "rule.criticality AS criticality, coalesce(rule.criticalityRank,0) AS criticalityRank"
    )
    cypher = (
        match_clause +
        f"WITH {group_by}, u, " + _USER_TYP_STATUS_CYPHER + " "
        f"WITH {group_by}, count(DISTINCT u) AS userCount"
        + (", collect({id:u.id, name:coalesce(u.name,''), typ:typ, status:status}) AS users" if detailed else "")
        + " "
        f"RETURN {return_fields}, userCount" + (", users" if detailed else "") + " "
        f"ORDER BY coalesce({group_by}.criticalityRank,0) DESC, userCount DESC"
    )
    return [dict(r) for r in s.run(cypher, ruleset=ruleset, runId=run_id)]


def _summary_export_csv(rows: list[dict]) -> str:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["id", "name", "criticality", "userCount"])
    for r in rows:
        w.writerow([r["id"], r.get("shortDescription") or r.get("description") or "",
                    r["criticality"], r["userCount"]])
    return "﻿" + buf.getvalue()


def _summary_export_xlsx(rows: list[dict], detailed: bool, sheet_title: str) -> bytes:
    """Baut die Ergebnisse-Uebersicht als natives Excel. detailed=True gruppiert je Zeile die
    Nutzerliste darunter ueber Excels Gliederungs-/Gruppierungsfunktion (eingeklappt, ueber das
    '+'-Symbol am linken Rand auffaltbar) statt sie ungruppiert in die Tabelle zu quetschen."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(["id", "name", "criticality", "userCount"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    if detailed:
        # summaryBelow=False: die Zeile der Query/Regel steht OBEN, die gruppierten Nutzerzeilen
        # folgen darunter (Standard waere umgekehrt, wie bei einer Summenzeile unter Detailzeilen).
        ws.sheet_properties.outlinePr.summaryBelow = False
    for r in rows:
        ws.append([r["id"], r.get("shortDescription") or r.get("description") or "",
                   r["criticality"], r["userCount"]])
        if detailed:
            users = r.get("users") or []
            if users:
                start = ws.max_row + 1
                ws.append(["", "id", "name", "typ", "status"])
                for cell in ws[ws.max_row]:
                    cell.font = Font(italic=True)
                for u in users:
                    ws.append(["", u.get("id"), u.get("name"), u.get("typ"), u.get("status")])
                ws.row_dimensions.group(start, ws.max_row, outline_level=1, hidden=True)
    widths = {"A": 16, "B": 44, "C": 14, "D": 12, "E": 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _summary_export_response(s, kind: str, run_id: str, fmt: str, filename_base: str, sheet_title: str):
    if fmt not in ("csv", "xlsx", "xlsx_detailed"):
        raise HTTPException(400, f"Unbekanntes Format '{fmt}' (csv|xlsx|xlsx_detailed erwartet)")
    detailed = fmt == "xlsx_detailed"
    rows = _summary_export_rows(s, kind, run_id, detailed)
    if fmt == "csv":
        return Response(content=_summary_export_csv(rows), media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": f'attachment; filename="{filename_base}_{run_id}.csv"'})
    try:
        data = _summary_export_xlsx(rows, detailed, sheet_title)
    except ImportError:
        raise HTTPException(500, "openpyxl nicht installiert — Image neu bauen: docker compose build backend")
    return Response(content=data,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename_base}_{run_id}.xlsx"'})


@app.get("/queries/summary/export")
def export_queries_summary(runId: str, format: str = "csv"):
    """Einzelfilter-Uebersicht (s. GET /queries/summary) als CSV/Excel/ausfuehrliches Excel
    (format=csv|xlsx|xlsx_detailed) — ausfuehrlich faltet je Query die betroffenen Nutzer per
    Excel-Gruppierung auf."""
    with driver.session() as s:
        return _summary_export_response(s, "query", runId, format, "einzelfilter_uebersicht", "Einzelfilter")


@app.get("/sodrules/summary/export")
def export_sod_rules_summary(runId: str, format: str = "csv"):
    """SoD-Regel-Uebersicht (s. GET /sodrules/summary) als CSV/Excel/ausfuehrliches Excel
    (format=csv|xlsx|xlsx_detailed) — ausfuehrlich faltet je Regel die betroffenen Nutzer per
    Excel-Gruppierung auf."""
    with driver.session() as s:
        return _summary_export_response(s, "sod", runId, format, "sod_uebersicht", "SoD-Regeln")


@app.get("/matches")
def matches(runId: str, query: str | None = None, user: str | None = None,
            userType: list[str] = Query(default=[])):
    """Wer matcht Query X (Einzelberechtigung) im Zwischenergebnis (:User)-[:MATCHES]->(:Query)
    eines Laufs — optional auf einen User/Nutzertyp(en) eingeschraenkt (Drill-down 'Query -> wer matcht')."""
    with driver.session() as s:
        return [dict(r) for r in s.run(
            "MATCH (u:User)-[:MATCHES {runId:$runId}]->(q:Query) "
            "WHERE ($qid IS NULL OR q.id = $qid) AND ($user IS NULL OR u.id = $user) "
            "  AND (size($userTypes) = 0 OR any(t IN $userTypes WHERE t IN labels(u))) "
            "RETURN u.id AS user, coalesce(u.name,'') AS name, "
            "       CASE WHEN 'Dialog' IN labels(u) THEN 'Dialog' WHEN 'System' IN labels(u) THEN 'System' "
            "            WHEN 'Service' IN labels(u) THEN 'Service' WHEN 'Communication' IN labels(u) THEN 'Comm' ELSE '?' END AS typ, "
            "       CASE WHEN 'Locked' IN labels(u) THEN 'gesperrt' ELSE 'aktiv' END AS status, "
            "       q.id AS query "
            "ORDER BY status, user", runId=runId, qid=query, user=user, userTypes=userType)]


_SATISFIED_BY_CYPHER = (
    "MATCH (u:User {id:$user, dataset:$dataset}) "
    "OPTIONAL MATCH (u)-[g:ASSIGNED_TO]->(roleActor:Role) "
    "  WHERE (g.validFrom IS NULL OR g.validFrom<=$asOf) AND (g.validTo IS NULL OR $asOf<=g.validTo) "
    "OPTIONAL MATCH (u)-[:HAS_PROFILE]->(profActor:Profile) "
    "WITH u, [x IN collect(DISTINCT roleActor) WHERE x IS NOT NULL] "
    "   + [x IN collect(DISTINCT profActor) WHERE x IS NOT NULL] AS actors "
    "UNWIND actors AS actor "
    "MATCH (actor)-[:CONTAINS|HAS_PROFILE*0..4]->(via)-[:HAS_AUTH]->(a:Authorization {dataset:$dataset, object:$object}) "
    "WHERE all(r IN $reqs WHERE "
    "  r.field IN $orgFields "
    "  OR ( apoc.any.property(a,'f_'+r.field) IS NOT NULL "
    "       AND ( '*' IN apoc.any.property(a,'f_'+r.field) "
    "             OR CASE WHEN r.andLogic "
    "                  THEN all(v IN r.values WHERE v IN apoc.any.property(a,'f_'+r.field) "
    "                         OR any(rg IN apoc.any.property(a,'f_'+r.field) WHERE rg CONTAINS '..' AND split(rg,'..')[0]<=v AND v<=split(rg,'..')[1])) "
    "                  ELSE any(v IN r.values WHERE v IN apoc.any.property(a,'f_'+r.field) "
    "                         OR any(rg IN apoc.any.property(a,'f_'+r.field) WHERE rg CONTAINS '..' AND split(rg,'..')[0]<=v AND v<=split(rg,'..')[1])) "
    "                END ) ) ) "
    "RETURN DISTINCT labels(actor)[0] AS actorType, actor.id AS actorId, "
    # "via" = der Knoten, von dem die tatsaechlich treffende Authorization ausgeht. Ist via der
    # Akteur selbst (Pfadlaenge 0), kommt der Treffer aus dessen EIGENER, gepflegter Definition
    # (z. B. Role-HAS_AUTH aus AGR_1251). Ist via ein anderer Knoten (z. B. das generierte Profil
    # der Rolle, oder eine ueber CONTAINS enthaltene Sammelrolle), kommt der Treffer aus DIESEM
    # Pfad -- mit ggf. abweichenden Werten zur eigenen Definition (Design vs. tatsaechlich
    # Generiertes, vgl. Konsistenzcheck D4 stale_profile_generation.cypher). Beide Faelle werden
    # hier NICHT zusammengefasst -- das Backend liefert alle Belege, die UI kennzeichnet die Quelle
    # je Zeile. WERTIDENTISCHE eigene-Definition/generiertes-Profil-Paare fasst das Frontend
    # (rcCollapseActors) zu einer Zeile zusammen; nur bei ABWEICHUNG bleiben beide (Divergenz -> D4).
    "  (CASE WHEN via = actor THEN null ELSE labels(via)[0] END) AS viaType, "
    "  (CASE WHEN via = actor THEN null ELSE via.id END) AS viaId, "
    # "technisch" = ein generiertes Profil (PFCG-Artefakt einer Rolle), keine eigenstaendig
    # gepflegte Berechtigung -- redundant zur ohnehin angezeigten Rolle. Bewusst NICHT auf die
    # Rollen DIESES Users beschraenkt (erste Version tat das und uebersah "verwaiste" generierte
    # Profile, deren erzeugende Rolle im Extrakt nicht mehr existiert, z. B. geloescht/umbenannt
    # -- Nutzer-Beispiel T-EC37002026): EXISTS { (:Role)-[:HAS_PROFILE]->(actor) } prueft, ob
    # IRGENDEINE Rolle dieses Profil erzeugt (im Graphen unabhaengig vom betrachteten User/
    # Zeitpunkt) -- Stichprobe sachsenenergie: erfasst 3737/3829 T-*-Profile strukturell sowie
    # 1171 generierte Profile ohne "T-"-Praefix. Der Praefix-Fallback faengt die restlichen
    # verwaisten T-*-Profile (92 von 3829) ab, deren erzeugende Rolle fehlt.
    "  (labels(actor)[0] = 'Profile' AND EXISTS { MATCH (:Role)-[:HAS_PROFILE]->(actor) }) AS hasGeneratingRole, "
    "  (labels(actor)[0] = 'Profile' AND ( "
    "    EXISTS { MATCH (:Role)-[:HAS_PROFILE]->(actor) } OR actor.id STARTS WITH 'T-' "
    "  )) AS technical, "
    "  [k IN keys(a) WHERE k STARTS WITH 'f_' | {field: substring(k,2), values: apoc.any.property(a,k)}] AS authFields "
    "ORDER BY actorType, actorId"
)


def _query_objects(s, ruleset: str, dataset: str, as_of, user: str, org_fields: list[str], query_id: str) -> list[dict]:
    """Objekt-/TCode-Aufschluesselung einer einzelnen Query fuer einen User (Kernlogik von
    Root-Cause) — wiederverwendet sowohl fuer den Einzelfilter- als auch den SoD-Regel-Aufruf."""
    qrec = s.run(
        "MATCH (q:Query {id:$qid, ruleset:$ruleset}) "
        "OPTIONAL MATCH (q)-[:REQUIRES]->(ar:AuthReq) "
        "RETURN q.tcodes AS tcodes, q.disregardTcode AS disregardTcode, "
        "  collect(CASE WHEN ar IS NULL THEN null ELSE "
        "    {object:ar.object, field:ar.field, andLogic:ar.andLogic, values:ar.values} END) AS reqs",
        qid=query_id, ruleset=ruleset).single()
    if qrec is None:
        raise HTTPException(404, f"Query '{query_id}' nicht gefunden")
    reqs = [r for r in qrec["reqs"] if r is not None]

    def satisfied_by(obj: str, obj_reqs: list[dict]) -> list[dict]:
        rows = s.run(_SATISFIED_BY_CYPHER, user=user, dataset=dataset, asOf=as_of,
                     object=obj, reqs=obj_reqs, orgFields=org_fields)
        # Ein Akteur kann denselben Treffer ueber MEHRERE Authorization-Knoten erreichen -- z. B.
        # eine Rolle ueber ihre eigene, gepflegte Definition (AGR_1251 -> Role-HAS_AUTH) UND
        # zusaetzlich ueber ihr generiertes Profil (Role-HAS_PROFILE->Profile-HAS_AUTH), mit ggf.
        # abweichenden Werten (Design vs. tatsaechlich Generiertes -- vgl. Konsistenzcheck D4,
        # stale_profile_generation.cypher). Bewusst NICHT zusammengefasst: beide sind
        # eigenstaendige Belege; "via" zeigt der UI, aus welcher Quelle eine Zeile stammt.
        return [{"actorType": r["actorType"], "actorId": r["actorId"], "technical": r["technical"],
                 "orphaned": bool(r["technical"]) and not bool(r["hasGeneratingRole"]),
                 "via": {"type": r["viaType"], "id": r["viaId"]} if r["viaId"] else None,
                 "authValues": [f"{f['field']}={','.join(f['values'])}" for f in r["authFields"]]}
                for r in rows]

    objects = sorted({r["object"] for r in reqs})
    result_objects = [{"object": obj, "requirement": [r for r in reqs if r["object"] == obj],
                        "satisfiedBy": satisfied_by(obj, [r for r in reqs if r["object"] == obj])}
                       for obj in objects]

    tcodes = qrec["tcodes"] or []
    disregard = bool(qrec["disregardTcode"])
    blocks = []
    if not disregard and tcodes and "*" not in tcodes:
        tcode_req = [{"field": "TCD", "andLogic": False, "values": tcodes}]
        blocks.append({"object": "S_TCODE (TCode-Prüfung)", "requirement": tcode_req,
                        "satisfiedBy": satisfied_by("S_TCODE", tcode_req)})
    blocks += result_objects
    return blocks


@app.get("/root-cause")
def root_cause(runId: str, user: str, query: str | None = None, rule: str | None = None):
    """Root-Cause-Erklaerung fuer einen User: entweder fuer eine einzelne Query (Einzelfilter) oder
    fuer eine SoD-Regel (alle Klauseln, je die vom User tatsaechlich gematchte(n) Query(s)). Pro
    Berechtigungsobjekt (+ TCode-Pruefung als eigener Pseudo-Block) wird gezeigt, WELCHE Anforderung
    galt und welche Rolle(n)/Profil(e) sie mit welcher konkreten Authorization erfuellen — macht
    sichtbar, wenn unterschiedliche Objekte/Klauseln durch unterschiedliche Rollen gedeckt werden
    (AE-03/AE-09), nicht nur 'welche Rolle' wie die Evidenz (VIA_ROLE/VIA_PROFILE, AE-11)."""
    if not query and not rule:
        raise HTTPException(400, "entweder 'query' oder 'rule' angeben")
    with driver.session() as s:
        run = s.run("MATCH (r:Run {runId:$id}) RETURN r.ruleset AS ruleset, r.dataset AS dataset, "
                    "r.asOf AS asOf", id=runId).single()
        if not run:
            raise HTTPException(404, f"Lauf '{runId}' nicht gefunden")
        ruleset, dataset, as_of = run["ruleset"], run["dataset"], run["asOf"]
        org_fields = [r["field"] for r in s.run(
            "MATCH (of:OrgField {dataset:$d}) RETURN of.field AS field", d=dataset)]

        if query:
            blocks = [{"label": f"Query {query}",
                       "objects": _query_objects(s, ruleset, dataset, as_of, user, org_fields, query)}]
        else:
            clauses = s.run(
                "MATCH (rule:SoDRule {id:$rid, ruleset:$ruleset})-[:HAS_CLAUSE]->(cl:Clause)-[:NEEDS]->(q:Query) "
                "RETURN cl.idx AS idx, collect(q.id) AS qids ORDER BY idx",
                rid=rule, ruleset=ruleset)
            blocks = []
            for c in clauses:
                idx, qids = c["idx"], c["qids"]
                matched = [r["qid"] for r in s.run(
                    "MATCH (u:User {id:$user})-[:MATCHES {runId:$runId}]->(q:Query) "
                    "WHERE q.id IN $qids RETURN q.id AS qid", user=user, runId=runId, qids=qids)]
                if not matched:
                    blocks.append({"label": f"Klausel {idx + 1}", "objects": [],
                                   "note": "keine gematchte Query in dieser Klausel gefunden"})
                    continue
                for qid in matched:
                    blocks.append({"label": f"Klausel {idx + 1} · Query {qid}",
                                   "objects": _query_objects(s, ruleset, dataset, as_of, user, org_fields, qid)})
            if not blocks:
                raise HTTPException(404, f"SoD-Regel '{rule}' nicht gefunden")

        return {"queryId": query, "ruleId": rule, "user": user, "blocks": blocks}


@app.get("/findings/export")
def export_findings(runId: str, minRank: int = 0, user: str | None = None, rule: str | None = None,
                     userType: list[str] = Query(default=[]), sleeping: str | None = None,
                     ruleCriticality: str | None = None, sleepDaysOverride: int | None = None,
                     locked: str | None = None, lockReason: str | None = None):
    """Findings eines Laufs als CSV (Semikolon, UTF-8-BOM -> Excel-tauglich) — Ergebnis-Export
    getrennt vom Quell-Backup. Nimmt dieselben Filter-Parameter wie GET /findings (identische
    _FINDINGS_WHERE) an, damit der Export standardmaessig zur gerade angezeigten (gefilterten)
    Tabelle passt statt immer den kompletten Lauf zu dumpen; ohne Parameter = alles (bisheriges
    Verhalten). Enthaelt jetzt auch die Regel-Bezeichnung (fehlte komplett, nur die ID war drin)."""
    with driver.session() as s:
        rows = list(s.run(
            "MATCH (u:User)-[:VIOLATES]->(f:SoDConflict {runId:$runId})-[:BASED_ON]->(rule:SoDRule) "
            "MATCH (run:Run {runId:$runId}) "
            + _FINDINGS_WHERE +
            "RETURN u.id AS user, f.ruleId AS rule, "
            "coalesce(rule.shortDescription, rule.description, '') AS ruleName, "
            "f.ruleset AS ruleset, f.dataset AS dataset, "
            "f.criticality AS criticality, coalesce(f.criticalityRank,0) AS criticalityRank, "
            "effSleeping AS sleeping, effLastLogonKnown AS lastLogonKnown, "
            "f.conflictType AS conflictType, "
            "[(f)-[:VIA_ROLE]->(r) | r.id] AS roles, [(f)-[:VIA_PROFILE]->(p) | p.id] AS profiles, "
            "f.reasonCode AS reasonCode "
            "ORDER BY criticalityRank DESC, user",
            runId=runId, minRank=minRank, user=user, rule=rule,
            userTypes=userType, sleeping=sleeping, ruleCriticality=ruleCriticality,
            sleepDaysOverride=sleepDaysOverride, locked=locked, lockReason=lockReason))
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["user", "rule", "ruleName", "ruleset", "dataset", "criticality", "criticalityRank",
                "sleeping", "lastLogonKnown", "conflictType", "roles", "profiles", "reasonCode"])
    for x in rows:
        w.writerow([x["user"], x["rule"], x["ruleName"], x["ruleset"], x["dataset"], x["criticality"],
                    x["criticalityRank"], x["sleeping"], x["lastLogonKnown"], x["conflictType"],
                    "|".join(x["roles"] or []), "|".join(x["profiles"] or []), x["reasonCode"]])
    data = "﻿" + buf.getvalue()   # BOM, damit Excel UTF-8/Umlaute korrekt liest
    return Response(content=data, media_type="text/csv; charset=utf-8",
                   headers={"Content-Disposition": f'attachment; filename="findings_{runId}.csv"'})


@app.get("/matches/export")
def export_matches(runId: str, query: str | None = None, user: str | None = None,
                    userType: list[str] = Query(default=[])):
    """Einzelfilter-Treffer ('wer matcht Query X') als CSV -- Pendant zu /findings/export fuer die
    Einzelfilter-Ansicht (Matches-Tabelle). Nimmt dieselben Parameter wie GET /matches an, damit
    der Export zur aktuell angezeigten Tabelle passt (der Export-Button exportierte bisher IMMER
    die SoD-Findings, auch waehrend die Matches-Tabelle sichtbar war)."""
    with driver.session() as s:
        rows = list(s.run(
            "MATCH (u:User)-[:MATCHES {runId:$runId}]->(q:Query) "
            "WHERE ($qid IS NULL OR q.id = $qid) AND ($user IS NULL OR u.id = $user) "
            "  AND (size($userTypes) = 0 OR any(t IN $userTypes WHERE t IN labels(u))) "
            "RETURN u.id AS user, coalesce(u.name,'') AS name, "
            "  CASE WHEN 'Dialog' IN labels(u) THEN 'Dialog' WHEN 'System' IN labels(u) THEN 'System' "
            "       WHEN 'Service' IN labels(u) THEN 'Service' WHEN 'Communication' IN labels(u) THEN 'Communication' "
            "       WHEN 'Reference' IN labels(u) THEN 'Reference' ELSE '?' END AS typ, "
            "  CASE WHEN 'Locked' IN labels(u) THEN 'gesperrt' ELSE 'aktiv' END AS status, "
            "  q.id AS query, coalesce(q.shortDescription, q.description, '') AS queryName, "
            "  coalesce(q.criticality,'') AS criticality "
            "ORDER BY status, user", runId=runId, qid=query, user=user, userTypes=userType))
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["user", "name", "typ", "status", "query", "queryName", "criticality"])
    for x in rows:
        w.writerow([x["user"], x["name"], x["typ"], x["status"], x["query"], x["queryName"], x["criticality"]])
    data = "﻿" + buf.getvalue()
    return Response(content=data, media_type="text/csv; charset=utf-8",
                   headers={"Content-Disposition": f'attachment; filename="matches_{runId}.csv"'})


@app.post("/imports")
def start_import(req: ImportReq):
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "import", "request": req.model_dump()}
    threading.Thread(target=do_import, args=(job_id, req), daemon=True).start()
    return {"jobId": job_id}


@app.post("/imports/upload")
async def upload_import(file: UploadFile = File(...), dataset: str = Form(""),
                        lang: str = Form(""), skipSchema: bool = Form(False),
                        asOf: str = Form(""), clearFirst: bool = Form(False)):
    """Import per ZIP-Upload: .csv/.txt im ZIP -> data/import/<dataset> -> Import-Job."""
    ds = re.sub(r"[^A-Za-z0-9._-]", "_", dataset or Path(file.filename or "import").stem)
    if not ds:
        raise HTTPException(400, "dataset-Name leer/ungueltig")
    tmp = Path(tempfile.gettempdir()) / f"upload_{uuid.uuid4().hex}.zip"
    with open(tmp, "wb") as out:
        shutil.copyfileobj(file.file, out)
    langs = [c.strip() for c in lang.split(",") if c.strip()]
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "import",
                    "request": {"dataset": ds, "upload": file.filename}}
    threading.Thread(target=do_upload_import,
                     args=(job_id, str(tmp), ds, langs, skipSchema, asOf or None, clearFirst),
                     daemon=True).start()
    return {"jobId": job_id, "dataset": ds}


@app.get("/datasets/{datasetId}/import-state")
def get_import_state(datasetId: str):
    """Import-Checkpoint lesen: welche Load-Schritte wurden schon abgeschlossen?"""
    return _read_state(datasetId) or {}


@app.get("/datasets/{datasetId}/run-state")
def get_run_state(datasetId: str):
    """Lauf-Checkpoint lesen: ist fuer dieses Dataset ein Lauf abgebrochen und fortsetzbar?
    (Einzel- und Batch-Laeufe teilen sich denselben Checkpoint -- pro Dataset ist immer nur ein
    Lauf gleichzeitig aktiv/fortsetzbar, wie beim Import.)"""
    return _read_checkpoint(_checkpoint_path(datasetId, "_run_state.json")) or {}


@app.post("/datasets/{datasetId}/run-state/discard")
def discard_run_state(datasetId: str):
    """Verwirft einen abgebrochenen Lauf-Checkpoint (nur die Datei) -- Aufraeumen der
    Graphdaten des abgebrochenen Laufs selbst erfolgt separat ueber POST /runs/{runId}/delete."""
    _clear_checkpoint(_checkpoint_path(datasetId, "_run_state.json"))
    return {"dataset": datasetId, "discarded": True}


@app.get("/import-folders")
def import_folders():
    """Vorhandene data/import/<dataset>-Ordner (mit/ohne bereits konvertierte .csv)."""
    if not DATA_DIR.exists():
        return []
    out = []
    for d in sorted(p for p in DATA_DIR.iterdir() if p.is_dir()):
        txt, csv = len(list(d.glob("*.txt"))), len(list(d.glob("*.csv")))
        if txt or csv:   # nur echte Import-Ordner (SE16-.txt oder bereits konvertierte .csv)
            out.append({"dataset": d.name, "txt": txt, "csv": csv})
    return out


class RunMetaReq(BaseModel):
    title: str | None = None          # leer -> runId als Fallback (analog RunReq.title)
    description: str | None = None    # freier, mehrzeiliger Text; leer -> geloescht


@app.put("/runs/{runId}/meta")
def set_run_meta(runId: str, req: RunMetaReq):
    """Titel/Beschreibung eines bestehenden Laufs nachtraeglich aendern -- reine Metadaten,
    kein Neu-Lauf (Nutzer-Feedback: Variantenname war nach dem Anlegen nicht mehr korrigierbar)."""
    title = (req.title or "").strip() or runId
    description = (req.description or "").strip() or None
    with driver.session() as s:
        rec = s.run("MATCH (r:Run {runId:$id}) SET r.title=$title, r.description=$description "
                    "RETURN r.title AS title, r.description AS description",
                    id=runId, title=title, description=description).single()
        if not rec:
            raise HTTPException(404, f"Run '{runId}' nicht gefunden")
    return {"runId": runId, "title": rec["title"], "description": rec["description"]}


@app.post("/runs/{runId}/explain")
def explain_run(runId: str):
    """Evidenz (verursachende Rollen/Profile, intra/inter) fuer einen Lauf nachrechnen (teuer)."""
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "explain", "request": {"runId": runId}}
    threading.Thread(target=do_explain, args=(job_id, runId), daemon=True).start()
    return {"jobId": job_id}


@app.post("/runs/{runId}/delete")
def delete_run(runId: str):
    """Loescht einen einzelnen Auswertungslauf (Run + Findings); Dataset bleibt unberuehrt."""
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "delete_run", "request": {"runId": runId}}
    threading.Thread(target=do_delete_run, args=(job_id, runId), daemon=True).start()
    return {"jobId": job_id}


@app.post("/runs/{runId}/backup")
def backup_run(runId: str):
    """Sichert einen Auswertungslauf (Run + Findings, ohne Evidenz) als eigenes ZIP."""
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "backup_run", "request": {"runId": runId}}
    threading.Thread(target=do_backup_run, args=(job_id, runId), daemon=True).start()
    return {"jobId": job_id}


@app.get("/runs/backups")
def list_run_backups():
    if not RUN_BACKUP_DIR.exists():
        return []
    out = []
    for p in sorted(RUN_BACKUP_DIR.glob("*.zip"), key=lambda x: x.stat().st_mtime, reverse=True):
        with zipfile.ZipFile(p) as z:
            manifest = json.loads(z.read("manifest.json"))
        out.append({"file": p.name, **manifest, "sizeBytes": p.stat().st_size})
    return out


@app.get("/runs/backups/{file}/download")
def download_run_backup(file: str):
    return FileResponse(_run_backup_path(file), filename=file, media_type="application/zip")


class RunRestoreReq(BaseModel):
    force: bool = False     # true = trotz abweichender Dataset-uid wiederherstellen


@app.post("/runs/backups/{file}/restore")
def restore_run_backup(file: str, req: RunRestoreReq):
    path = _run_backup_path(file)
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "restore_run", "request": {"file": file, "force": req.force}}
    threading.Thread(target=do_restore_run, args=(job_id, path, req.force), daemon=True).start()
    return {"jobId": job_id}


@app.post("/datasets/{dataset}/clear")
def clear_dataset(dataset: str):
    """Loescht ein Dataset (inkl. Runs/Findings); Ruleset + Schema bleiben."""
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "clear", "request": {"dataset": dataset}}
    threading.Thread(target=do_clear, args=(job_id, dataset), daemon=True).start()
    return {"jobId": job_id}


@app.post("/reset")
def reset_data():
    """Setzt alle Daten zurueck (alle Datasets/Runs/Findings); Ruleset + Schema bleiben."""
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "reset", "request": {}}
    threading.Thread(target=do_reset, args=(job_id,), daemon=True).start()
    return {"jobId": job_id}


@app.post("/datasets/{dataset}/backup")
def backup_dataset(dataset: str, clear: bool = False):
    """Sichert die .csv eines Datasets als ZIP; clear=true leert danach den Graph (Backup & Clear)."""
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "backup", "request": {"dataset": dataset, "clear": clear}}
    threading.Thread(target=do_backup, args=(job_id, dataset, clear), daemon=True).start()
    return {"jobId": job_id}


@app.get("/backups")
def list_backups():
    if not BACKUP_DIR.exists():
        return []
    out = []
    for p in sorted(BACKUP_DIR.glob("*.zip"), key=lambda x: x.stat().st_mtime, reverse=True):
        st = p.stat()
        out.append({"file": p.name, "dataset": p.name.split("__")[0], "sizeBytes": st.st_size,
                    "createdAt": datetime.datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds")})
    return out


@app.get("/backups/{file}/download")
def download_backup(file: str):
    return FileResponse(_backup_path(file), filename=file, media_type="application/zip")


class RestoreReq(BaseModel):
    dataset: str | None = None      # Ziel-Dataset (Default: aus dem Manifest des Backups)


@app.post("/backups/{file}/restore")
def restore_backup(file: str, req: RestoreReq):
    path = _backup_path(file)
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "restore", "request": {"file": file, "dataset": req.dataset}}
    threading.Thread(target=do_restore, args=(job_id, path, req.dataset), daemon=True).start()
    return {"jobId": job_id}


@app.post("/runs")
def start_run(req: RunReq):
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "run", "request": req.model_dump()}
    threading.Thread(target=do_run, args=(job_id, req), daemon=True).start()
    return {"jobId": job_id}


@app.post("/runs/batch")
def start_run_batch(req: RunBatchReq):
    """Mehrere Org-Varianten in einem Schritt anlegen — je Variante ein eigener (:Run) mit dem
    Variantennamen als Titel (s. ROADMAP.md, 'Multi-Varianten-Laeufe'). Ein Job, sequenziell
    abgearbeitet (gemeinsame Neo4j-Session, kein paralleles Schreiben auf dasselbe Dataset)."""
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "kind": "run_batch", "request": req.model_dump()}
    threading.Thread(target=do_run_batch, args=(job_id, req), daemon=True).start()
    return {"jobId": job_id}


@app.get("/jobs/{job_id}")
def job_status(job_id: str):
    if job_id not in jobs:
        raise HTTPException(404, "unbekannte Job-Id")
    return jobs[job_id]


@app.post("/jobs/{job_id}/cancel")
def cancel_job(job_id: str):
    """Setzt cancelRequested-Flag; der laufende Thread bricht nach dem naechsten Schritt ab."""
    if job_id not in jobs:
        raise HTTPException(404, "unbekannte Job-Id")
    if jobs[job_id].get("status") not in ("queued", "running"):
        raise HTTPException(409, "Job ist nicht mehr aktiv")
    jobs[job_id]["cancelRequested"] = True
    return {"jobId": job_id}


# Minimale UI (Phase 9, Bau-Schritt 2): statische Single-Page, vom Backend ausgeliefert.
# Bewusst leichtgewichtig (kein Node/React-Build, ein Container) — das gebrandete Frontend mit
# Cytoscape.js-Graph bleibt der spaetere "Fancy"-Schritt. MUSS nach allen API-Routen gemountet
# werden, damit "/" die UI liefert, ohne /health, /runs, ... zu ueberdecken.
if FRONTEND_DIR.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="ui")
