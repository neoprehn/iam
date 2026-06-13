#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Einmalige Migration: KPMG-Ruleset (5 Excel-Tabellen) -> normalisierte, wartbare JSON.

NICHT Teil der Runtime-Pipeline (die bleibt container-only). Reines Dev-/Konvertierungs-
Werkzeug: nach der Migration ist die JSON das gepflegte Artefakt (oder eine spaetere
Editor-Ausbaustufe schreibt sie). Ausgabe: queries.json, sod_rules.json, legends.json,
ruleset.json im selben Ordner (UTF-8, ensure_ascii=False -> Umlaute bleiben erhalten).

Aufruf:  python rules/KPMG/_convert_kpmg.py
"""
import json, os, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))           # Quell-Ordner (Archiv): Excel + dieses Skript
OUT = os.path.normpath(os.path.join(HERE, "..", "..", "KPMG_R3"))  # gepflegtes JSON-Ruleset
os.makedirs(OUT, exist_ok=True)

def sheet_rows(path, sheet="Sheet1"):
    wb = openpyxl.load_workbook(os.path.join(HERE, path), read_only=True, data_only=True)
    ws = [w for w in wb.worksheets if w.title == sheet][0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({hdr[i]: (r[i] if i < len(r) else None) for i in range(len(hdr))})
    return out

def s(v):
    return "" if v is None else str(v).strip()

def b(v):
    return s(v).lower() == "true"

def split_values(v):
    # "02, 03" -> ["02","03"]; '*' und Einzelwerte bleiben erhalten
    t = s(v)
    if t == "":
        return []
    return [p.strip() for p in t.split(",") if p.strip() != ""]

# ---- Queries: Master (Spine) + Berechtigungen + Transaktionen ---------------------
queries = {}
for row in sheet_rows("Table view Query.xlsx"):
    name = s(row.get("Query name"))
    if not name:
        continue
    queries[name] = {
        "query": name,
        "description": s(row.get("Description")),
        "sortOrder": s(row.get("Sorting")),
        "useNaming": b(row.get("Use naming")),
        "soxClassification": s(row.get("Sox Classification")),
        "gdprClassification": s(row.get("Gdpr Classification")),
        "type": s(row.get("Type")),
        "multipleRun": b(row.get("Multiple run")),
        "disregardTcode": b(row.get("Disregard T-code")),
        "auditSave": s(row.get("Audit save")),
        "authorizations": [],
        "transactions": [],
    }

def ensure(name, desc=""):
    if name not in queries:
        queries[name] = {"query": name, "description": desc, "sortOrder": "", "useNaming": False,
                         "soxClassification": "", "gdprClassification": "", "type": "",
                         "multipleRun": False, "disregardTcode": False, "auditSave": "",
                         "authorizations": [], "transactions": []}
    return queries[name]

for row in sheet_rows("Query authorizations overview.xlsx"):
    name = s(row.get("Query name"))
    if not name:
        continue
    # Object-/Field-Texte bewusst NICHT uebernommen: Objekttexte stehen im Graph
    # (AuthObject.text aus TOBJT/13), Feldtexte sind SAP-Metadaten (kein Ruleset-Logikinhalt).
    # Konsistent zu den CSI-Rulesets, die diese Beschreibungen ebenfalls nicht fuehren.
    ensure(name, s(row.get("Query description")))["authorizations"].append({
        "object": s(row.get("Object")),
        "field": s(row.get("Field")),
        "andLogic": b(row.get("And-logic")),
        "values": split_values(row.get("Values")),
        "audit": b(row.get("Audit")),
    })

for row in sheet_rows("Query transaction relation.xlsx"):
    name = s(row.get("Query name"))
    if not name:
        continue
    # TCode-Text bewusst NICHT uebernommen: steht im Graph (Transaction.text aus TSTCT/09).
    ensure(name, s(row.get("Query description")))["transactions"].append({
        "tcode": s(row.get("Transaction")),
        "audit": b(row.get("Audit")),
        "stad": b(row.get("STAD")),
    })

queries_list = sorted(queries.values(), key=lambda q: q["query"])

# ---- SoD-Regeln: Master + Variablen->Query-Mapping --------------------------------
rules = {}
for row in sheet_rows("Table view SoD rule.xlsx"):
    rid = s(row.get("Text"))            # 'Text' enthaelt die Regel-ID (z. B. BCX_0001)
    if not rid:
        continue
    rules[rid] = {
        "sodRule": rid,
        "reasonCode": s(row.get("Reason code")),
        "description": s(row.get("Description")),
        "definition": s(row.get("Definition")),
        "definitionDescription": s(row.get("Def. description")),
        "expression": s(row.get("Expression")),
        "variables": {},                # Variable -> {query, description, type}
    }

for row in sheet_rows("Table view SoD rule detail with query info.xlsx"):
    rid = s(row.get("SoD rule"))
    var = s(row.get("Variable"))
    if not rid or not var:
        continue
    r = rules.setdefault(rid, {"sodRule": rid, "reasonCode": s(row.get("Reason code")),
                               "description": s(row.get("SoD description")), "definition": "",
                               "definitionDescription": "", "expression": s(row.get("Expression")),
                               "variables": {}})
    r["variables"][var] = {
        "query": s(row.get("Query")),
        "queryDescription": s(row.get("Query description")),
        "queryType": s(row.get("Query type")),
    }

rules_list = sorted(rules.values(), key=lambda r: r["sodRule"])

# ---- Legenden aus den ValueList_Helper-Blaettern ----------------------------------
def helper(path):
    wb = openpyxl.load_workbook(os.path.join(HERE, path), read_only=True, data_only=True)
    ws = [w for w in wb.worksheets if "ValueList" in w.title]
    rows = list(ws[0].iter_rows(values_only=True)) if ws else []
    wb.close()
    return rows

qh = helper("Table view Query.xlsx")          # col1 misc, col2 sox-scale, col3 type-code, col4 type-desc
types = {}
sox_scale = []
for r in qh:
    if len(r) > 3 and r[2] is not None:
        types[s(r[2])] = s(r[3])
    if len(r) > 1 and r[1] is not None:
        sox_scale.append(s(r[1]))
dh = helper("Table view SoD rule detail with query info.xlsx")  # col1 reason, col2 expr-templates
reason_codes = [s(r[0]) for r in dh if r and r[0] is not None]
expr_templates = [s(r[1]) for r in dh if len(r) > 1 and r[1] is not None]

legends = {
    "soxClassificationScale": [x for x in sox_scale if x],
    "queryTypes_toolLegend": {k: v for k, v in types.items() if k},
    "typeUsage": {
        "JAP": "Jahresabschluss-relevant (Pflicht-Scope)",
        "AO": "optionale Query",
        "TEST": "Test-Query",
    },
    "reasonCodes": [x for x in reason_codes if x],
    "expressionTemplates": [x for x in expr_templates if x],
    "_note": "typeUsage = fachliche Bedeutung der genutzten type-Codes (vom Mandanten geklaert): type ist ein "
             "SCOPE-/Auswahl-Filter, KEIN Verknuepfungsoperator. queryTypes_toolLegend = Original-Text aus den "
             "ValueList_Helper-Blaettern (weicht ab, nur als Referenz).",
}

ruleset = {
    "ruleset": "kpmg_r3",
    "name": "KPMG SoD-Ruleset (R/3)",
    "platform": "R/3",
    "note": "Noch nicht fuer S/4 geeignet (Fiori/OData-Ebene, geaenderte TCodes/Objekte fehlen).",
    "model": "query-based (Queries = Funktionsbausteine; SoD-Regeln = boolesche Ausdruecke ueber Query-Variablen)",
    "source": "rules/_archive/KPMG/*.xlsx (5 Tabellen)",
    "generatedAt": datetime.date.today().isoformat(),
    "combinationSemantics": {
        "valuesWithinField": "AND wenn andLogic=true, sonst OR (Spalte G der Excel)",
        "fieldsWithinObject": "AND",
        "objectsWithinQuery": "AND",
        "transactions": "OR (mehrere TCodes); tcode '*' = beliebig -> TCode-Teil trivial erfuellt",
        "authVsTransaction": "AND (User braucht passenden TCode UND die Berechtigungsobjekte); fehlt ein Teil ganz, zaehlt nur der vorhandene",
        "type": "SCOPE-Filter (welche Queries laufen), KEIN Operator: JAP=Jahresabschluss, AO=optional, TEST=Test",
        "_note": "Vom Mandanten bestaetigtes Auswerte-Modell (Basis fuer den Phase-3-Evaluator). Beispiel 1003: (SNRO OR SNUM) AND (S_NUMBER.ACTVT=11|13) AND (S_NUMBER.NROBJ=*).",
    },
    "counts": {
        "queries": len(queries_list),
        "queriesWithAuthorizations": sum(1 for q in queries_list if q["authorizations"]),
        "queriesWithTransactions": sum(1 for q in queries_list if q["transactions"]),
        "sodRules": len(rules_list),
        "variableMappings": sum(len(r["variables"]) for r in rules_list),
    },
}

def dump(name, obj):
    p = os.path.join(OUT, name)
    with open(p, "w", encoding="utf-8") as fh:
        json.dump(obj, fh, ensure_ascii=False, indent=2)
    return os.path.getsize(p)

sizes = {
    "queries.json": dump("queries.json", queries_list),
    "sod_rules.json": dump("sod_rules.json", rules_list),
    "legends.json": dump("legends.json", legends),
    "ruleset.json": dump("ruleset.json", ruleset),
}

print("=== geschrieben ===")
for k, v in sizes.items():
    print(f"  {k}: {v:,} Bytes")
print("=== ruleset.counts ===")
print(json.dumps(ruleset["counts"], indent=2, ensure_ascii=False))
# Encoding-Probe: ein Wert mit Umlaut (repr, um U+FFFD-Korruption auszuschliessen)
probe = next((r["description"] for r in rules_list if any(ch in r["description"] for ch in "äöüÄÖÜß") or "�" in r["description"]), "")
print("=== Encoding-Probe (Regel-Beschreibung mit Umlaut) ===")
print("  repr:", repr(probe))
